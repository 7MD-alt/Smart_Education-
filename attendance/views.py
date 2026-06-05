from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.exceptions import PermissionDenied
from rest_framework.parsers import MultiPartParser, FormParser
import os
import io
import re
import json
import logging
import requests as http_requests
from django.http import HttpResponse, StreamingHttpResponse
from django.core.mail import send_mail
from django.utils import timezone
from django.conf import settings as django_settings
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


# ── Student password generator ────────────────────────────────────────────────
import unicodedata, re

def generate_student_password(first_name: str, massar_code: str) -> str:
    """
    Auto-generate the initial password for a student account.
    Rule: lowercase-normalised first name + digits extracted from the massar code.
    E.g.  first_name="Ahmed",  massar_code="G123456789"  →  "ahmed123456789"
          first_name="Émilie", massar_code="K98765"      →  "emilie98765"
    Students log in with this password and can change it from their profile.
    """
    nfkd      = unicodedata.normalize("NFKD", first_name or "")
    name_part = re.sub(r"[^a-zA-Z]", "", nfkd).lower()
    digits    = re.sub(r"\D", "", massar_code or "")
    return f"{name_part}{digits}" or "student1234"

from .models import (
    User,
    AdminProfile,
    TeacherProfile,
    StudentProfile,
    Department,
    Filiere,
    Course,
    FiliereCourse,
    CourseMaterial,
    MaterialEmbedding,
    AttendanceRecord,
    ChatSession,
    ChatMessage,
    Notification,
    NotificationType,
    FaceRegistrationRequest,
    FaceRequestStatus,
    Seance,
    SeanceStatus,
    SessionType,
    TPGroup,
    generate_seance_code,
)
from datetime import date, datetime, timedelta

from .serializers import (
    UserSerializer,
    AdminProfileSerializer,
    TeacherProfileSerializer,
    StudentProfileSerializer,
    DepartmentSerializer,
    FiliereSerializer,
    CourseSerializer,
    FiliereCourseSerializer,
    CourseMaterialSerializer,
    MaterialEmbeddingSerializer,
    AttendanceRecordSerializer,
    ChatSessionSerializer,
    ChatMessageSerializer,
)

from .permissions import (
    IsAdminUserRole,
    IsTeacherUserRole,
    IsStudentUserRole,
    IsAdminOrTeacher,
)

from .services.face_recognition_service import recognize_and_mark_attendance
from .services.multi_agent_service import ask_tutor
from .services.face_registration_service import register_student_face
# campuseye_agent_service is imported lazily inside AgentExecuteAPIView to avoid circular imports
from .services.novaa_tutor_service import ask_novaa, ask_novaa_stream
from .services.platform_executor import get_platform_context, get_email_context_for_student


class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all().order_by("id")
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated, IsAdminUserRole]
    pagination_class = None   # return all users — admin panel handles filtering


class AdminProfileViewSet(viewsets.ModelViewSet):
    queryset = AdminProfile.objects.all()
    serializer_class = AdminProfileSerializer
    permission_classes = [IsAuthenticated, IsAdminUserRole]


class TeacherProfileViewSet(viewsets.ModelViewSet):
    queryset = TeacherProfile.objects.all()
    serializer_class = TeacherProfileSerializer
    permission_classes = [IsAuthenticated, IsAdminOrTeacher]
    pagination_class = None


class StudentProfileViewSet(viewsets.ModelViewSet):
    queryset = StudentProfile.objects.all()
    serializer_class = StudentProfileSerializer
    # Students access their own profile via /api/me/profile/ — this endpoint is admin/teacher only
    permission_classes = [IsAuthenticated, IsAdminOrTeacher]
    pagination_class = None

    @action(detail=True, methods=["get"])
    def attendance(self, request, pk=None):
        student = self.get_object()
        records = AttendanceRecord.objects.filter(student=student).select_related(
            "student", "student__user", "course"
        )
        serializer = AttendanceRecordSerializer(records, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=["get"])
    def chat_sessions(self, request, pk=None):
        student = self.get_object()
        sessions = ChatSession.objects.filter(student=student).select_related(
            "student", "student__user"
        )
        serializer = ChatSessionSerializer(sessions, many=True)
        return Response(serializer.data)


class DepartmentViewSet(viewsets.ModelViewSet):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer
    permission_classes = [IsAuthenticated, IsAdminUserRole]


class FiliereViewSet(viewsets.ModelViewSet):
    queryset = Filiere.objects.all()
    serializer_class = FiliereSerializer
    permission_classes = [IsAuthenticated, IsAdminUserRole]

    @action(detail=True, methods=["get"])
    def courses(self, request, pk=None):
        filiere = self.get_object()
        links = FiliereCourse.objects.filter(filiere=filiere).select_related(
            "course", "course__teacher", "course__teacher__user"
        )
        serializer = FiliereCourseSerializer(links, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=["get"])
    def students(self, request, pk=None):
        filiere = self.get_object()
        students = StudentProfile.objects.filter(filiere=filiere).select_related(
            "user", "filiere", "filiere__department"
        )
        serializer = StudentProfileSerializer(students, many=True)
        return Response(serializer.data)


class CourseViewSet(viewsets.ModelViewSet):
    queryset = Course.objects.all()
    serializer_class = CourseSerializer
    permission_classes = [IsAuthenticated, IsAdminOrTeacher]

    def get_permissions(self):
        if self.request.method in ("GET", "HEAD", "OPTIONS"):
            return [IsAuthenticated()]
        return [IsAuthenticated(), IsAdminOrTeacher()]

    def _is_owner_or_admin(self, course):
        user = self.request.user
        if user.role == "ADMIN":
            return True
        try:
            return course.teacher.user == user
        except Exception:
            return False

    def perform_create(self, serializer):
        course = serializer.save()
        # ── Notify teacher they've been assigned a new course ────────────
        try:
            if course.teacher:
                _push_notification(
                    user=course.teacher.user,
                    notif_type=NotificationType.COURSE_ASSIGNED,
                    title="You were assigned to a new course",
                    message=f'You have been assigned to teach "{course.title}".',
                    link="/teacher",
                    metadata={"course_id": course.id},
                )
        except Exception as exc:
            logger.error(f"[Notif] course create trigger failed: {exc}")

    def perform_update(self, serializer):
        old_teacher_id = self.get_object().teacher_id if self.get_object().teacher else None
        course = serializer.save()
        # ── Notify teacher if they were newly assigned (teacher changed) ─
        try:
            new_teacher = course.teacher
            if new_teacher and new_teacher.id != old_teacher_id:
                _push_notification(
                    user=new_teacher.user,
                    notif_type=NotificationType.COURSE_ASSIGNED,
                    title="You were assigned to a new course",
                    message=f'You have been assigned to teach "{course.title}".',
                    link="/teacher",
                    metadata={"course_id": course.id},
                )
        except Exception as exc:
            logger.error(f"[Notif] course update trigger failed: {exc}")

    def update(self, request, *args, **kwargs):
        if not self._is_owner_or_admin(self.get_object()):
            return Response({"error": "You can only edit your own courses."}, status=status.HTTP_403_FORBIDDEN)
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if not self._is_owner_or_admin(self.get_object()):
            return Response({"error": "You can only delete your own courses."}, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=["get"])
    def materials(self, request, pk=None):
        course = self.get_object()
        materials = CourseMaterial.objects.filter(course=course).select_related("course")
        serializer = CourseMaterialSerializer(materials, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=["get"])
    def attendance(self, request, pk=None):
        course = self.get_object()
        records = AttendanceRecord.objects.filter(course=course).select_related(
            "course", "student", "student__user"
        )
        serializer = AttendanceRecordSerializer(records, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=["get"], url_path="danger-zone-students")
    def danger_zone_students(self, request, pk=None):
        course = self.get_object()

        filieres = Filiere.objects.filter(filiere_courses__course=course).distinct()

        students = StudentProfile.objects.filter(
            filiere__in=filieres
        ).select_related("user", "filiere").distinct()

        results = []

        for student in students:
            absences = AttendanceRecord.objects.filter(
                student=student,
                course=course,
                status="ABSENT"
            ).count()

            max_absences = course.max_absences

            if absences >= max_absences - 1:
                results.append({
                    "student_id": student.student_id,
                    "full_name": f"{student.user.first_name} {student.user.last_name}".strip(),
                    "absences": absences,
                    "max_absences": max_absences,
                    "status": "DANGER" if absences >= max_absences else "WARNING",
                })

        return Response({
            "course_id": course.id,
            "course_title": course.title,
            "danger_students": results
        })


class FiliereCourseViewSet(viewsets.ModelViewSet):
    queryset = FiliereCourse.objects.all()
    serializer_class = FiliereCourseSerializer
    permission_classes = [IsAuthenticated, IsAdminOrTeacher]

    def perform_create(self, serializer):
        fc = serializer.save()
        # ── Notify the teacher that their course has a new filière/students ──
        try:
            course   = fc.course
            filiere  = fc.filiere
            teacher  = course.teacher
            if teacher:
                students_count = StudentProfile.objects.filter(filiere=filiere).count()
                _push_notification(
                    user=teacher.user,
                    notif_type=NotificationType.STUDENT_JOINED,
                    title=f"New filière linked to {course.title}",
                    message=(
                        f'Filière "{filiere.name}" ({students_count} student{"s" if students_count != 1 else ""}) '
                        f"has been linked to your course {course.title}."
                    ),
                    link=f"/teacher",
                    metadata={"course_id": course.id, "filiere_id": filiere.id},
                )
        except Exception as exc:
            logger.error(f"[Notif] filiere-course create trigger failed: {exc}")


class CourseMaterialViewSet(viewsets.ModelViewSet):
    queryset = CourseMaterial.objects.all()
    serializer_class = CourseMaterialSerializer
    permission_classes = [IsAuthenticated, IsAdminOrTeacher]

    def perform_create(self, serializer):
        course = serializer.validated_data["course"]

        if (
            self.request.user.role == "TEACHER"
            and course.teacher.user != self.request.user
        ):
            raise PermissionDenied("You can only upload materials to your own courses.")

        material = serializer.save()

        # ── Notify all enrolled students ──────────────────────────────────
        try:
            filiere_ids = FiliereCourse.objects.filter(course=course).values_list("filiere_id", flat=True)
            students = StudentProfile.objects.filter(filiere_id__in=filiere_ids).select_related("user")
            mat_name = material.file.name.split("/")[-1] if material.file else "a file"
            for student in students:
                _push_notification(
                    user=student.user,
                    notif_type=NotificationType.MATERIAL_ADDED,
                    title=f"New material in {course.title}",
                    message=f'Your teacher uploaded "{mat_name}" to {course.title}.',
                    link=f"/student/courses/{course.id}/materials",
                    metadata={"course_id": course.id, "material_id": material.id},
                )
        except Exception as exc:
            logger.error(f"[Notif] material upload trigger failed: {exc}")

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()

        if (
            request.user.role == "TEACHER"
            and instance.course.teacher_id != request.user.pk
        ):
            return Response(
                {"error": "You can only delete materials from your own courses."},
                status=status.HTTP_403_FORBIDDEN,
            )

        return super().destroy(request, *args, **kwargs)


class MaterialEmbeddingViewSet(viewsets.ModelViewSet):
    queryset = MaterialEmbedding.objects.all()
    serializer_class = MaterialEmbeddingSerializer
    permission_classes = [IsAuthenticated, IsAdminOrTeacher]


class AttendanceRecordViewSet(viewsets.ModelViewSet):
    serializer_class = AttendanceRecordSerializer

    def get_permissions(self):
        """
        - Safe methods (GET/HEAD/OPTIONS): any authenticated user
        - Writes (POST/PUT/PATCH/DELETE): admin or teacher only
        """
        if self.request.method in ("POST", "PUT", "PATCH", "DELETE"):
            return [IsAuthenticated(), IsAdminOrTeacher()]
        return [IsAuthenticated()]

    def get_queryset(self):
        user = self.request.user
        if user.role == "STUDENT":
            try:
                return AttendanceRecord.objects.filter(
                    student=user.student_profile
                ).select_related("course", "course__teacher", "course__teacher__user")
            except Exception:
                return AttendanceRecord.objects.none()
        # Admin / Teacher: full access
        return AttendanceRecord.objects.all().select_related(
            "student", "student__user", "course", "course__teacher"
        )


class ChatSessionViewSet(viewsets.ModelViewSet):
    serializer_class = ChatSessionSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        try:
            student = user.student_profile
            return ChatSession.objects.filter(student=student).order_by("-started_at")
        except Exception:
            return ChatSession.objects.none()

    def perform_create(self, serializer):
        """Auto-assign the student from the logged-in user — ignore any student_id from the client."""
        try:
            student = self.request.user.student_profile
        except Exception:
            from rest_framework.exceptions import ValidationError
            raise ValidationError("No student profile found for this user.")
        serializer.save(student=student)


class ChatMessageViewSet(viewsets.ModelViewSet):
    serializer_class = ChatMessageSerializer
    permission_classes = [IsAuthenticated, IsStudentUserRole]

    def get_queryset(self):
        try:
            student = self.request.user.student_profile
        except Exception:
            return ChatMessage.objects.none()

        qs = ChatMessage.objects.filter(
            session__student=student
        ).select_related("session")

        session_id = self.request.query_params.get("session")
        if session_id:
            qs = qs.filter(session__id=session_id)

        return qs.order_by("timestamp")


class AttendanceScanAPIView(APIView):
    permission_classes = [IsAuthenticated, IsTeacherUserRole]

    def post(self, request):
        image = request.FILES.get("image")
        course_id = request.data.get("course_id")

        if not image or not course_id:
            return Response(
                {"error": "image and course_id are required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            course = Course.objects.select_related("teacher").get(pk=course_id)
        except Course.DoesNotExist:
            return Response(
                {"error": "Course not found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        if course.teacher_id != request.user.pk:
            return Response(
                {"error": "You can only scan attendance for your own courses."},
                status=status.HTTP_403_FORBIDDEN,
            )

        result = recognize_and_mark_attendance(image, course_id)

        if result.get("success"):
            # ── Build full class roster for today ──────────────────────────
            today = date.today()
            filiere_ids = FiliereCourse.objects.filter(course=course).values_list("filiere_id", flat=True)
            enrolled = (
                StudentProfile.objects
                .filter(filiere_id__in=filiere_ids)
                .select_related("user")
                .order_by("user__last_name", "user__first_name")
            )
            today_records = {
                r.student_id: r.status
                for r in AttendanceRecord.objects.filter(course=course, date=today)
            }

            recognized_ids = {s["student_id"] for s in result.get("recognized_students", [])}

            roster = []
            for s in enrolled:
                rec_status = today_records.get(s.user_id)
                roster.append({
                    "student_id": s.student_id,
                    "full_name":  f"{s.user.first_name} {s.user.last_name}".strip() or s.user.username,
                    "status":     rec_status if rec_status else "ABSENT",
                    "just_recognized": s.student_id in recognized_ids,
                })

            result["roster"]       = roster
            result["total_enrolled"] = len(roster)
            result["present_count"]  = sum(1 for r in roster if r["status"] == "PRESENT")
            result["absent_count"]   = sum(1 for r in roster if r["status"] == "ABSENT")
            result["late_count"]     = sum(1 for r in roster if r["status"] == "LATE")
            result["scan_date"]      = str(today)
            result["course_title"]   = course.title
            return Response(result, status=status.HTTP_200_OK)

        return Response(result, status=status.HTTP_400_BAD_REQUEST)


class ChatAskAPIView(APIView):
    permission_classes = [IsAuthenticated, IsStudentUserRole]

    def post(self, request):
        question  = request.data.get("question")
        course_id = request.data.get("course_id")
        student_id = request.data.get("student_id") or request.user.id
        mode      = request.data.get("mode")  # optional forced agent label

        if not question:
            return Response(
                {"error": "question is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        file_context = request.data.get("file_context")
        result = ask_tutor(
            question,
            student_id=student_id,
            course_id=course_id,
            mode=mode,
            file_context=file_context,
        )
        return Response(result, status=status.HTTP_200_OK)


# ── Supported plain-text / code extensions ────────────────────────────────────
_TEXT_EXTS = {
    ".py", ".js", ".ts", ".jsx", ".tsx", ".java", ".c", ".cpp", ".cs",
    ".go", ".rs", ".php", ".rb", ".swift", ".kt", ".html", ".css",
    ".sql", ".sh", ".bash", ".json", ".xml", ".yaml", ".yml",
    ".txt", ".md", ".csv",
}
_MAX_FILE_MB = 5
_MAX_CONTEXT_CHARS = 12_000   # ~3 k tokens — safe for all Groq models


class ChatFileUploadAPIView(APIView):
    """
    POST /api/chat/upload/
    Accepts a single file (PDF, DOCX, or any text/code file).
    Extracts its text and returns it so the frontend can pass it
    as `file_context` on the next chat/ask/ request.
    """
    permission_classes = [IsAuthenticated, IsStudentUserRole]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        file = request.FILES.get("file")
        if not file:
            return Response({"error": "No file provided."}, status=status.HTTP_400_BAD_REQUEST)

        if file.size > _MAX_FILE_MB * 1024 * 1024:
            return Response(
                {"error": f"File is too large. Maximum allowed size is {_MAX_FILE_MB} MB."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        filename = file.name
        ext = os.path.splitext(filename)[1].lower()
        text = ""

        try:
            if ext == ".pdf":
                import pdfplumber, io
                with pdfplumber.open(io.BytesIO(file.read())) as pdf:
                    pages = [p.extract_text() or "" for p in pdf.pages]
                text = "\n\n".join(pages)

            elif ext == ".docx":
                import docx, io
                doc = docx.Document(io.BytesIO(file.read()))
                text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())

            elif ext in _TEXT_EXTS:
                raw = file.read()
                text = raw.decode("utf-8", errors="replace")

            else:
                return Response(
                    {"error": f"Unsupported file type '{ext}'. "
                               "Please upload a PDF, DOCX, or a code/text file."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        except Exception as exc:
            logger.error("File extraction failed for '%s': %s", filename, exc)
            return Response(
                {"error": f"Could not read '{filename}'. Make sure it is a valid file."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        text = text.strip()
        if not text:
            return Response(
                {"error": "The file appears to be empty or has no extractable text."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        truncated = len(text) > _MAX_CONTEXT_CHARS
        return Response({
            "success":   True,
            "filename":  filename,
            "text":      text[:_MAX_CONTEXT_CHARS],
            "preview":   text[:300],
            "truncated": truncated,
        }, status=status.HTTP_200_OK)


class StudentRegisterFaceAPIView(APIView):
    permission_classes = [IsAuthenticated, IsAdminOrTeacher]

    def post(self, request):
        student_id = request.data.get("student_id")
        image = request.FILES.get("image")

        if not student_id or not image:
            return Response(
                {"error": "student_id and image are required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        result = register_student_face(student_id, image)

        if result.get("success"):
            return Response(result, status=status.HTTP_200_OK)

        return Response(result, status=status.HTTP_400_BAD_REQUEST)


class AdminStatsAPIView(APIView):
    permission_classes = [IsAuthenticated, IsAdminUserRole]

    def get(self, request):
        data = {
            "users": User.objects.count(),
            "students": StudentProfile.objects.count(),
            "teachers": TeacherProfile.objects.count(),
            "departments": Department.objects.count(),
            "filieres": Filiere.objects.count(),
            "courses": Course.objects.count(),
            "materials": CourseMaterial.objects.count(),
        }
        return Response(data)


class AdminStudentDetailAPIView(APIView):
    """Return full info about one student for the admin view."""
    permission_classes = [IsAuthenticated, IsAdminUserRole]

    def get(self, request, user_id):
        try:
            profile = StudentProfile.objects.select_related("user", "filiere").get(user__id=user_id)
        except StudentProfile.DoesNotExist:
            return Response({"error": "Student not found."}, status=404)

        user = profile.user

        # Courses the student is enrolled in (via FiliereCourse matching filiere + semester)
        filiere_courses = FiliereCourse.objects.filter(
            filiere=profile.filiere,
            semester=profile.semester,
        ).select_related("course", "course__teacher", "course__teacher__user")

        courses_data = []
        total_present = total_absent = total_late = 0

        for fc in filiere_courses:
            course = fc.course
            records = AttendanceRecord.objects.filter(student=profile, course=course)
            present = records.filter(status="PRESENT").count()
            absent  = records.filter(status="ABSENT").count()
            late    = records.filter(status="LATE").count()
            total   = records.count()

            total_present += present
            total_absent  += absent
            total_late    += late

            pct = round((present / total * 100), 1) if total > 0 else None

            courses_data.append({
                "id":           course.id,
                "title":        course.title,
                "teacher":      f"{course.teacher.user.first_name} {course.teacher.user.last_name}".strip() if course.teacher else "—",
                "max_absences": course.max_absences,
                "present":      present,
                "absent":       absent,
                "late":         late,
                "total":        total,
                "attendance_pct": pct,
                "status": (
                    "DANGER"  if absent >= course.max_absences else
                    "WARNING" if absent >= course.max_absences - 1 else
                    "OK"
                ),
            })

        total_records = total_present + total_absent + total_late
        overall_pct = round(total_present / total_records * 100, 1) if total_records > 0 else None

        data = {
            "id":         user.id,
            "username":   user.username,
            "email":      user.email,
            "first_name": user.first_name,
            "last_name":  user.last_name,
            "is_active":  user.is_active,
            "student_id": profile.student_id,
            "semester":   profile.semester,
            "has_face":   bool(profile.face_encoding),
            "filiere": {
                "id":   profile.filiere.id   if profile.filiere else None,
                "name": profile.filiere.name if profile.filiere else "—",
                "code": profile.filiere.code if profile.filiere else "—",
            },
            "stats": {
                "total_courses":  len(courses_data),
                "total_present":  total_present,
                "total_absent":   total_absent,
                "total_late":     total_late,
                "total_records":  total_records,
                "attendance_pct": overall_pct,
            },
            "courses": courses_data,
        }
        return Response(data)


class AdminTeacherDetailAPIView(APIView):
    """Return full info about one teacher for the admin view."""
    permission_classes = [IsAuthenticated, IsAdminUserRole]

    def get(self, request, user_id):
        try:
            profile = TeacherProfile.objects.select_related("user", "department").get(user__id=user_id)
        except TeacherProfile.DoesNotExist:
            return Response({"error": "Teacher not found."}, status=404)

        user = profile.user

        courses = Course.objects.filter(teacher=profile).prefetch_related("filiere_courses")

        courses_data = []
        total_records = total_present = total_absent = total_late = 0

        for course in courses:
            records = AttendanceRecord.objects.filter(course=course)
            present = records.filter(status="PRESENT").count()
            absent  = records.filter(status="ABSENT").count()
            late    = records.filter(status="LATE").count()
            total   = records.count()
            students = StudentProfile.objects.filter(
                filiere__filiere_courses__course=course
            ).distinct().count()

            total_records += total
            total_present += present
            total_absent  += absent
            total_late    += late

            # Materials count
            materials = CourseMaterial.objects.filter(course=course).count()

            courses_data.append({
                "id":           course.id,
                "title":        course.title,
                "max_absences": course.max_absences,
                "students":     students,
                "materials":    materials,
                "present":      present,
                "absent":       absent,
                "late":         late,
                "total_records": total,
                "attendance_pct": round(present / total * 100, 1) if total > 0 else None,
            })

        data = {
            "id":         user.id,
            "username":   user.username,
            "email":      user.email,
            "first_name": user.first_name,
            "last_name":  user.last_name,
            "is_active":  user.is_active,
            "department": {
                "id":   profile.department.id   if profile.department else None,
                "name": profile.department.name if profile.department else "N/A",
                "code": profile.department.code if profile.department else "N/A",
            },
            "stats": {
                "total_courses":   len(courses_data),
                "total_materials": sum(c["materials"] for c in courses_data),
                "total_students":  sum(c["students"]  for c in courses_data),
                "total_records":   total_records,
                "total_present":   total_present,
                "total_absent":    total_absent,
                "total_late":      total_late,
                "attendance_pct":  round(total_present / total_records * 100, 1) if total_records > 0 else None,
            },
            "courses": courses_data,
        }
        return Response(data)


class AdminImportUsersAPIView(APIView):
    """
    POST /api/admin/import-users/
    Accepts a multipart CSV file. Creates User + role profile for each row.
    CSV columns: username, role, first_name, last_name, email,
                 password (optional for STUDENT — auto-generated from first_name+massar digits),
                 student_id, massar_code, filiere_id, semester  (STUDENT),
                 department_id (TEACHER/ADMIN)
    Returns: { created, skipped, errors: [{row, reason}] }
    """
    permission_classes = [IsAuthenticated, IsAdminUserRole]
    parser_classes = [MultiPartParser, FormParser]

    REQUIRED      = {"username", "role"}
    VALID_ROLES   = {"STUDENT", "TEACHER", "ADMIN"}

    def post(self, request):
        import csv, io as _io
        from django.db import transaction

        csv_file = request.FILES.get("file")
        if not csv_file:
            return Response({"error": "No file provided."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            decoded = csv_file.read().decode("utf-8-sig")
            reader = csv.DictReader(_io.StringIO(decoded))
        except Exception as exc:
            return Response({"error": f"Could not read CSV: {exc}"}, status=status.HTTP_400_BAD_REQUEST)

        created = 0
        skipped = 0
        errors  = []
        base_student_count   = StudentProfile.objects.count()
        students_in_import   = 0

        for row_num, row in enumerate(reader, start=2):   # row 1 = header
            row = {k.strip().lower(): (v or "").strip() for k, v in row.items()}

            # ── Validate required fields ───────────────────────────────
            missing = [f for f in self.REQUIRED if not row.get(f)]
            if missing:
                errors.append({"row": row_num, "reason": f"Missing required fields: {', '.join(missing)}"})
                continue

            role = row["role"].upper()
            if role not in self.VALID_ROLES:
                errors.append({"row": row_num, "reason": f"Invalid role '{row['role']}'. Must be STUDENT, TEACHER, or ADMIN."})
                continue

            # Non-student rows still need an explicit password
            if role != "STUDENT" and not row.get("password"):
                errors.append({"row": row_num, "reason": "password is required for TEACHER and ADMIN rows."})
                continue

            username = row["username"]
            if User.objects.filter(username=username).exists():
                errors.append({"row": row_num, "reason": f"Username '{username}' already exists."})
                skipped += 1
                continue

            try:
                with transaction.atomic():
                    first_name  = row.get("first_name", "")
                    massar_code = row.get("massar_code", "").strip().upper() or None

                    # Auto-generate password for students if not provided
                    if role == "STUDENT":
                        password = row.get("password") or generate_student_password(first_name, massar_code or "")
                    else:
                        password = row["password"]

                    user = User.objects.create_user(
                        username=username,
                        password=password,
                        first_name=first_name,
                        last_name=row.get("last_name", ""),
                        email=row.get("email", ""),
                        role=role,
                    )

                    if role == "STUDENT":
                        filiere = None
                        if row.get("filiere_id"):
                            try:
                                filiere = Filiere.objects.get(pk=int(row["filiere_id"]))
                            except (Filiere.DoesNotExist, ValueError):
                                pass
                        semester = int(row.get("semester") or 1)
                        auto_sid = str(base_student_count + students_in_import + 1)
                        StudentProfile.objects.create(
                            user=user,
                            student_id=row.get("student_id") or auto_sid,
                            massar_code=massar_code,
                            filiere=filiere,
                            semester=semester,
                        )
                        students_in_import += 1

                    elif role == "TEACHER":
                        dept = None
                        if row.get("department_id"):
                            try:
                                dept = Department.objects.get(pk=int(row["department_id"]))
                            except (Department.DoesNotExist, ValueError):
                                pass
                        if dept is None:
                            dept = Department.objects.first()
                        if dept is None:
                            raise ValueError("No department found. Create a department first.")
                        TeacherProfile.objects.create(user=user, department=dept)

                    elif role == "ADMIN":
                        dept = None
                        if row.get("department_id"):
                            try:
                                dept = Department.objects.get(pk=int(row["department_id"]))
                            except (Department.DoesNotExist, ValueError):
                                pass
                        AdminProfile.objects.create(user=user, department=dept)

                    created += 1

            except Exception as exc:
                errors.append({"row": row_num, "reason": str(exc)})

        return Response({
            "created": created,
            "skipped": skipped,
            "errors":  errors,
            "total_processed": created + skipped + len(errors) - skipped,
        })


class TeacherStatsAPIView(APIView):
    permission_classes = [IsAuthenticated, IsTeacherUserRole]

    def get(self, request):
        teacher = request.user.teacher_profile
        courses = Course.objects.filter(teacher=teacher)

        students = StudentProfile.objects.filter(
            filiere__in=FiliereCourse.objects.filter(
                course__in=courses
            ).values_list("filiere", flat=True)
        ).distinct()

        data = {
            "courses": courses.count(),
            "materials": CourseMaterial.objects.filter(course__in=courses).count(),
            "students": students.count(),
            "attendance_records": AttendanceRecord.objects.filter(course__in=courses).count(),
        }
        return Response(data)


class StudentStatsAPIView(APIView):
    permission_classes = [IsAuthenticated, IsStudentUserRole]

    def get(self, request):
        student = request.user.student_profile

        courses = Course.objects.filter(
            filiere_courses__filiere=student.filiere,
            filiere_courses__semester=student.semester
        ).distinct()

        records = AttendanceRecord.objects.filter(student=student)

        data = {
            "courses": courses.count(),
            "attendance_records": records.count(),
            "absences": records.filter(status="ABSENT").count(),
            "chat_sessions": ChatSession.objects.filter(student=student).count(),
        }
        return Response(data)


class TeacherAttendanceSummaryAPIView(APIView):
    permission_classes = [IsAuthenticated, IsTeacherUserRole]

    def get(self, request):
        teacher = request.user.teacher_profile
        records = AttendanceRecord.objects.filter(course__teacher=teacher)

        total_records = records.count()
        present = records.filter(status="PRESENT").count()
        absent = records.filter(status="ABSENT").count()
        late = records.filter(status="LATE").count()

        attendance_rate = 0
        if total_records > 0:
            attendance_rate = round((present / total_records) * 100, 2)

        data = {
            "total_records": total_records,
            "present": present,
            "absent": absent,
            "late": late,
            "attendance_rate": attendance_rate,
        }
        return Response(data)


class StudentAttendanceSummaryAPIView(APIView):
    permission_classes = [IsAuthenticated, IsStudentUserRole]

    def get(self, request):
        student = request.user.student_profile
        records = AttendanceRecord.objects.filter(student=student)

        total_records = records.count()
        present = records.filter(status="PRESENT").count()
        absent = records.filter(status="ABSENT").count()
        late = records.filter(status="LATE").count()

        attendance_rate = 0
        if total_records > 0:
            attendance_rate = round((present / total_records) * 100, 2)

        danger_courses = []
        courses = Course.objects.filter(
            attendance_records__student=student
        ).distinct()

        for course in courses:
            student_absences = AttendanceRecord.objects.filter(
                student=student,
                course=course,
                status="ABSENT"
            ).count()

            if student_absences >= course.max_absences - 1:
                danger_courses.append({
                    "course_id": course.id,
                    "course_title": course.title,
                    "absences": student_absences,
                    "max_absences": course.max_absences,
                    "status": "DANGER" if student_absences >= course.max_absences else "WARNING",
                })

        data = {
            "total_records": total_records,
            "present": present,
            "absent": absent,
            "late": late,
            "attendance_rate": attendance_rate,
            "danger_courses": danger_courses,
        }
        return Response(data)


class MeAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        data = {
            "id": user.id,
            "username": user.username,
            "email": user.email,
            "first_name": user.first_name,
            "last_name": user.last_name,
            "role": user.role,
            "is_active": user.is_active,
        }
        return Response(data)

    def patch(self, request):
        user = request.user
        data = request.data
        errors = {}

        # ── username change ──────────────────────────────────────────────────
        new_username = data.get("username", "").strip()
        if new_username and new_username != user.username:
            if User.objects.filter(username=new_username).exclude(pk=user.pk).exists():
                errors["username"] = "That username is already taken."
            else:
                user.username = new_username

        # ── password change ──────────────────────────────────────────────────
        current_password = data.get("current_password", "")
        new_password = data.get("new_password", "")

        if new_password:
            if not current_password:
                errors["current_password"] = "Current password is required to set a new one."
            elif not user.check_password(current_password):
                errors["current_password"] = "Current password is incorrect."
            elif len(new_password) < 8:
                errors["new_password"] = "New password must be at least 8 characters."
            else:
                user.set_password(new_password)

        if errors:
            return Response(errors, status=status.HTTP_400_BAD_REQUEST)

        user.save()
        return Response({
            "success": True,
            "username": user.username,
        })


class MeProfileAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user

        if user.role == "ADMIN":
            profile = AdminProfile.objects.filter(user=user).first()
            if not profile:
                return Response({"error": "Admin profile not found."}, status=status.HTTP_404_NOT_FOUND)
            return Response(AdminProfileSerializer(profile).data)

        if user.role == "TEACHER":
            profile = TeacherProfile.objects.filter(user=user).first()
            if not profile:
                return Response({"error": "Teacher profile not found."}, status=status.HTTP_404_NOT_FOUND)
            return Response(TeacherProfileSerializer(profile).data)

        if user.role == "STUDENT":
            profile = StudentProfile.objects.filter(user=user).first()
            if not profile:
                return Response({"error": "Student profile not found."}, status=status.HTTP_404_NOT_FOUND)
            return Response(StudentProfileSerializer(profile).data)

        return Response({"error": "Unknown role."}, status=status.HTTP_400_BAD_REQUEST)


class MeAttendanceAPIView(APIView):
    """
    GET /api/me/attendance/
    Returns the logged-in student's attendance records.
    Optional query param: ?course_id=<id>  → filter by course
    """
    permission_classes = [IsAuthenticated, IsStudentUserRole]

    def get(self, request):
        try:
            student = request.user.student_profile
        except StudentProfile.DoesNotExist:
            return Response({"error": "Student profile not found."}, status=status.HTTP_404_NOT_FOUND)

        records = AttendanceRecord.objects.filter(student=student).select_related(
            "course", "course__teacher", "course__teacher__user"
        )

        course_id = request.query_params.get("course_id")
        if course_id:
            records = records.filter(course__id=course_id)

        records = records.order_by("-date", "-timestamp")
        serializer = AttendanceRecordSerializer(records, many=True)
        return Response(serializer.data)


class StudentSelfRegisterFaceAPIView(APIView):
    """
    POST /api/me/register-face/
    Student submits a face photo for admin review. Does NOT register the face
    immediately — creates a FaceRegistrationRequest with status=PENDING.

    GET /api/me/register-face/
    Returns the student's latest face request (status + image URL).
    """
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def get(self, request):
        user = request.user
        if user.role != "STUDENT":
            return Response({"error": "Students only."}, status=status.HTTP_403_FORBIDDEN)
        try:
            student_profile = user.student_profile
        except Exception:
            return Response({"error": "Student profile not found."}, status=status.HTTP_404_NOT_FOUND)

        latest = FaceRegistrationRequest.objects.filter(student=student_profile).first()
        if not latest:
            return Response({"status": None})

        image_url = request.build_absolute_uri(latest.image.url) if latest.image else None
        return Response({
            "id":            latest.id,
            "status":        latest.status,
            "created_at":    latest.created_at.isoformat(),
            "reviewed_at":   latest.reviewed_at.isoformat() if latest.reviewed_at else None,
            "reject_reason": latest.reject_reason,
            "image_url":     image_url,
        })

    def post(self, request):
        user = request.user
        if user.role != "STUDENT":
            return Response({"error": "Only students can submit face requests."}, status=status.HTTP_403_FORBIDDEN)

        try:
            student_profile = user.student_profile
        except Exception:
            return Response({"error": "Student profile not found."}, status=status.HTTP_404_NOT_FOUND)

        # Block if there's already a pending request
        if FaceRegistrationRequest.objects.filter(student=student_profile, status=FaceRequestStatus.PENDING).exists():
            return Response(
                {"error": "You already have a pending face request. Please wait for admin review."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        image = request.FILES.get("image")
        if not image:
            return Response({"error": "image file is required."}, status=status.HTTP_400_BAD_REQUEST)

        face_req = FaceRegistrationRequest.objects.create(
            student=student_profile,
            image=image,
            status=FaceRequestStatus.PENDING,
        )

        # Notify all admins
        from .models import AdminProfile
        admin_users = User.objects.filter(role="ADMIN")
        full_name = f"{user.first_name} {user.last_name}".strip() or user.username
        for admin_user in admin_users:
            _push_notification(
                admin_user,
                "FACE_REQUEST",
                "New Face Registration Request",
                f"{full_name} (ID: {student_profile.student_id}) submitted a face photo for review.",
                link="/admin/face-requests",
            )

        return Response({
            "success": True,
            "request_id": face_req.id,
            "message": "Your request has been submitted and is pending admin review.",
        }, status=status.HTTP_200_OK)


# ─────────────────────────────────────────────────────────────────────────────
# Admin — Face Registration Request Review
# ─────────────────────────────────────────────────────────────────────────────

class AdminFaceRequestListAPIView(APIView):
    """
    GET /api/admin/face-requests/
    Returns all face registration requests (optionally filter by ?status=PENDING).
    """
    permission_classes = [IsAuthenticated, IsAdminUserRole]

    def get(self, request):
        qs = FaceRegistrationRequest.objects.select_related(
            "student__user", "reviewed_by"
        ).all()

        status_filter = request.query_params.get("status")
        if status_filter:
            qs = qs.filter(status=status_filter.upper())

        data = []
        for r in qs:
            image_url = request.build_absolute_uri(r.image.url) if r.image else None
            s = r.student
            u = s.user
            data.append({
                "id":            r.id,
                "status":        r.status,
                "created_at":    r.created_at.isoformat(),
                "reviewed_at":   r.reviewed_at.isoformat() if r.reviewed_at else None,
                "reject_reason": r.reject_reason,
                "image_url":     image_url,
                "student": {
                    "id":         s.pk,
                    "student_id": s.student_id,
                    "full_name":  f"{u.first_name} {u.last_name}".strip() or u.username,
                    "username":   u.username,
                    "email":      u.email,
                    "filiere":    s.filiere.code if s.filiere else None,
                    "semester":   s.semester,
                },
                "reviewed_by": r.reviewed_by.username if r.reviewed_by else None,
            })

        return Response({"requests": data, "count": len(data)})


class AdminFaceRequestActionAPIView(APIView):
    """
    POST /api/admin/face-requests/<id>/approve/
    POST /api/admin/face-requests/<id>/reject/
    """
    permission_classes = [IsAuthenticated, IsAdminUserRole]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, req_id, action):
        try:
            face_req = FaceRegistrationRequest.objects.select_related("student__user").get(pk=req_id)
        except FaceRegistrationRequest.DoesNotExist:
            return Response({"error": "Request not found."}, status=status.HTTP_404_NOT_FOUND)

        if face_req.status != FaceRequestStatus.PENDING:
            return Response({"error": "This request has already been reviewed."}, status=status.HTTP_400_BAD_REQUEST)

        if action == "approve":
            # Actually register the face now
            result = register_student_face(face_req.student.student_id, face_req.image)
            if not result.get("success"):
                return Response({"error": result.get("error", "Face registration failed.")}, status=status.HTTP_400_BAD_REQUEST)

            face_req.status      = FaceRequestStatus.APPROVED
            face_req.reviewed_at = timezone.now()
            face_req.reviewed_by = request.user
            face_req.save()

            # Notify the student
            _push_notification(
                face_req.student.user,
                "FACE_REQUEST",
                "Face Registration Approved",
                "Your face photo has been approved. You'll now be recognised automatically during attendance scans.",
                link="/student/profile",
            )
            return Response({"success": True, "message": "Face registered and request approved."})

        elif action == "reject":
            reason = request.data.get("reason", "").strip()
            face_req.status        = FaceRequestStatus.REJECTED
            face_req.reviewed_at   = timezone.now()
            face_req.reviewed_by   = request.user
            face_req.reject_reason = reason
            face_req.save()

            # Notify the student
            _push_notification(
                face_req.student.user,
                "FACE_REQUEST",
                "Face Registration Rejected",
                f"Your face photo was rejected.{(' Reason: ' + reason) if reason else ' Please resubmit a clearer photo.'}",
                link="/student/profile",
            )
            return Response({"success": True, "message": "Request rejected."})

        return Response({"error": "Invalid action. Use 'approve' or 'reject'."}, status=status.HTTP_400_BAD_REQUEST)


class MeCoursesAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user

        if user.role == "TEACHER":
            try:
                teacher_profile = user.teacher_profile
            except TeacherProfile.DoesNotExist:
                return Response({"error": "Teacher profile not found."}, status=status.HTTP_404_NOT_FOUND)

            courses = Course.objects.filter(
                teacher=teacher_profile
            ).select_related("teacher", "teacher__user").prefetch_related("filiere_courses__filiere")

            data = []
            for c in courses:
                # Filières linked to this course
                filiere_courses = c.filiere_courses.all()
                filiere_ids     = [fc.filiere_id for fc in filiere_courses]
                filiere_names   = [fc.filiere.code for fc in filiere_courses if fc.filiere]

                # Count enrolled students across all linked filières
                student_count = StudentProfile.objects.filter(
                    filiere_id__in=filiere_ids
                ).count()

                # Count uploaded materials
                material_count = CourseMaterial.objects.filter(course=c).count()

                serialized = CourseSerializer(c).data
                serialized["student_count"]  = student_count
                serialized["material_count"] = material_count
                serialized["filiere_names"]  = filiere_names
                data.append(serialized)

            return Response(data)

        if user.role == "STUDENT":
            try:
                student_profile = user.student_profile
            except StudentProfile.DoesNotExist:
                return Response({"error": "Student profile not found."}, status=status.HTTP_404_NOT_FOUND)

            courses = Course.objects.filter(
                filiere_courses__filiere=student_profile.filiere,
                filiere_courses__semester=student_profile.semester
            ).select_related("teacher", "teacher__user").distinct()

            # Annotate each course with material_count so the NOVAA
            # chat widget knows which courses have RAG-able content
            data = []
            for c in courses:
                serialized = CourseSerializer(c).data
                serialized["material_count"] = CourseMaterial.objects.filter(course=c).count()
                data.append(serialized)
            return Response(data)

        if user.role == "ADMIN":
            courses = Course.objects.select_related("teacher", "teacher__user").all()
            return Response(CourseSerializer(courses, many=True).data)

        return Response({"error": "Unknown role."}, status=status.HTTP_400_BAD_REQUEST)


# ══════════════════════════════════════════════════════════════
# PLATFORM ASSISTANT  (AI widget — works for all roles)
# ══════════════════════════════════════════════════════════════

PLATFORM_SYSTEM_PROMPT = """
You are the Smart Education Platform Assistant — a friendly guide embedded inside the Smart Education academic platform used by a Moroccan engineering school.

Your job is to help ADMINS, TEACHERS, and STUDENTS understand and use the platform. Always be concise and clear. Respond in the same language the user writes in (French, English, or Darija are all fine).

ROLES:
- ADMIN: manages users, departments, filieres, courses
- TEACHER: manages courses, uploads materials, scans attendance via face recognition, views danger zone
- STUDENT: views courses, tracks attendance, uses the AI Tutor

ADMIN FEATURES:
- Dashboard: stats overview
- Users page: create/edit/delete users. Student accounts include a face registration wizard.
- Departments page: manage academic departments
- Filieres page: manage programs linked to departments
- Courses page: manage courses, assign teachers, set max absences

TEACHER FEATURES:
- Dashboard: stats on courses, students, attendance rate
- Live Scan: select a course, camera scans faces every 2 seconds and marks students PRESENT automatically
- Course Materials: upload PDF/DOCX/TXT files per course — these feed the AI Tutor
- Danger Zone: see students at risk of exceeding their absence limit (WARNING or DANGER status)

STUDENT FEATURES:
- Dashboard: overview of courses, absences, attendance stats
- Profile page: personal info (student ID, filiere, semester)
- AI Tutor: multi-agent AI that summarizes materials, explains concepts, debugs code, generates quizzes, helps with research, creates study plans, drafts essays, and translates content

KEY CONCEPTS:
- DANGER ZONE: each course has a max_absences limit (default 3). Students who reach it show as WARNING or DANGER.
- FACE RECOGNITION: used for login and automatic attendance marking
- ACADEMIC STRUCTURE: Department → Filiere → Course. Students belong to a filiere + semester.
- ATTENDANCE STATUSES: PRESENT, ABSENT, LATE

NAVIGATION:
- Student sidebar: Dashboard · Profile · AI Tutor
- Teacher sidebar: Dashboard · Live Scan · Profile
- Admin sidebar: Dashboard · Users · Departments · Filieres · Courses

RULES:
- Keep answers short unless the user asks for detail
- If something is outside the platform, say so politely
- Never make up features that don't exist
- Suggest which page to navigate to when relevant
"""


class CourseStudentsAPIView(APIView):
    """
    GET /api/teacher/courses/<course_id>/students/
    Returns all students enrolled in the course's filière.
    Also returns any existing attendance records for a given date
    if ?date=YYYY-MM-DD is provided.
    """
    permission_classes = [IsAuthenticated, IsAdminOrTeacher]

    def get(self, request, course_id):
        try:
            course = Course.objects.select_related("teacher__user").get(pk=course_id)
        except Course.DoesNotExist:
            return Response({"error": "Course not found."}, status=status.HTTP_404_NOT_FOUND)

        # Get all filieres this course belongs to
        filiere_ids = FiliereCourse.objects.filter(course=course).values_list("filiere_id", flat=True)
        students = (
            StudentProfile.objects
            .filter(filiere_id__in=filiere_ids)
            .select_related("user")
            .order_by("user__last_name", "user__first_name")
        )

        # Check existing records for a specific date
        date_str = request.query_params.get("date")
        existing = {}
        if date_str:
            records = AttendanceRecord.objects.filter(course=course, date=date_str)
            existing = {r.student_id: r.status for r in records}

        data = []
        for s in students:
            data.append({
                "student_profile_id": s.user_id,
                "student_id":         s.student_id,
                "full_name":          f"{s.user.last_name} {s.user.first_name}".strip() or s.user.username,
                "status":             existing.get(s.user_id, "ABSENT"),
            })

        return Response({"students": data, "course_title": course.title})


# ── Email helper ─────────────────────────────────────────────────────────────

def _get_student_absence_count(student, course):
    """Return how many ABSENT records this student has in this course."""
    return AttendanceRecord.objects.filter(
        course=course, student=student, status="ABSENT"
    ).count()


def _classify_status(absences, max_absences):
    """Return 'DANGER', 'WARNING', or 'OK'."""
    if absences >= max_absences:
        return "DANGER"
    if absences >= max_absences - 1:
        return "WARNING"
    return "OK"


def _handle_absence_thresholds(student, course, absences_before, absences_after):
    """
    Called after every attendance record save.
    - Sends email + in-app notification when a new WARNING or DANGER threshold is crossed.
    - Deactivates the student account when DANGER is reached for the first time.
    Pass absences_before / absences_after so we only act when a threshold is *newly* crossed.
    """
    max_abs = course.max_absences
    status_before = _classify_status(absences_before, max_abs)
    status_after  = _classify_status(absences_after,  max_abs)

    thresholds = ["OK", "WARNING", "DANGER"]
    if thresholds.index(status_after) <= thresholds.index(status_before):
        return  # no new threshold crossed — nothing to do

    # ── Send email alert ──────────────────────────────────────────────────────
    _send_absence_alert(student.user, course, absences_after, max_abs, status_after)

    # ── In-app notification ───────────────────────────────────────────────────
    notif_type_map = {
        "WARNING": NotificationType.ABSENCE_WARNING,
        "DANGER":  NotificationType.ABSENCE_DANGER,
    }
    notif_msgs = {
        "WARNING": (
            f"You are approaching the absence limit for {course.title} "
            f"({absences_after}/{max_abs} absences)."
        ),
        "DANGER": (
            f"You have reached the maximum absences for {course.title} "
            f"({absences_after}/{max_abs}). Your account has been suspended — "
            f"contact the administration immediately."
        ),
    }
    _push_notification(
        user=student.user,
        notif_type=notif_type_map[status_after],
        title=f"{'⚠️ Warning' if status_after == 'WARNING' else '🚨 Danger'} — {course.title}",
        message=notif_msgs[status_after],
        link="/student/attendance",
        metadata={"course_id": course.id},
    )

    # ── Auto-deactivate on DANGER ─────────────────────────────────────────────
    if status_after == "DANGER" and student.user.is_active:
        student.user.is_active = False
        student.user.save(update_fields=["is_active"])
        logger.info(
            f"[CampusEye] Deactivated account for student {student.student_id} "
            f"after reaching {absences_after} absences in '{course.title}'."
        )


def _send_absence_alert(student_user, course, absences, max_absences, alert_status):
    """
    Fire a single alert email to the student.
    Silently logs on failure so it never crashes the save endpoint.
    """
    email = student_user.email
    if not email:
        return

    frontend_url = getattr(django_settings, "CAMPUSEYE_FRONTEND_URL", "http://localhost:5173")
    name = student_user.get_full_name() or student_user.username
    remaining = max_absences - absences

    if alert_status == "DANGER":
        subject = f"[CampusEye] DANGER — Absence limit reached in {course.title}"
        body = (
            f"Dear {name},\n\n"
            f"You have reached or exceeded the maximum number of absences allowed "
            f"in the course \"{course.title}\".\n\n"
            f"  Absences recorded : {absences}\n"
            f"  Maximum allowed   : {max_absences}\n\n"
            f"Your situation requires immediate administrative action. "
            f"Please contact your teacher or the academic office as soon as possible.\n\n"
            f"You can review your attendance on the CampusEye platform:\n"
            f"{frontend_url}/student\n\n"
            f"— CampusEye Academic Platform"
        )
    else:  # WARNING
        subject = f"[CampusEye] Warning — Approaching absence limit in {course.title}"
        body = (
            f"Dear {name},\n\n"
            f"This is a friendly reminder that you are approaching the maximum number "
            f"of absences allowed in the course \"{course.title}\".\n\n"
            f"  Absences recorded : {absences}\n"
            f"  Maximum allowed   : {max_absences}\n"
            f"  Remaining         : {remaining} absence{'s' if remaining != 1 else ''}\n\n"
            f"Please make sure to attend upcoming sessions to avoid exceeding the limit.\n\n"
            f"You can review your attendance on the CampusEye platform:\n"
            f"{frontend_url}/student\n\n"
            f"— CampusEye Academic Platform"
        )

    try:
        send_mail(
            subject=subject,
            message=body,
            from_email=django_settings.DEFAULT_FROM_EMAIL,
            recipient_list=[email],
            fail_silently=False,
        )
        logger.info(f"[CampusEye] Sent {alert_status} alert to {email} ({course.title})")
    except Exception as exc:
        logger.error(f"[CampusEye] Failed to send alert to {email}: {exc}")


class ManualAttendanceSaveAPIView(APIView):
    """
    POST /api/teacher/courses/<course_id>/attendance/save/
    Body: { date: "YYYY-MM-DD", records: [{student_profile_id, status}, ...] }
    Creates or updates attendance records. Auto-sends email alerts when a
    student crosses into WARNING or DANGER territory.
    """
    permission_classes = [IsAuthenticated, IsAdminOrTeacher]

    def post(self, request, course_id):
        try:
            course = Course.objects.select_related("teacher").get(pk=course_id)
        except Course.DoesNotExist:
            return Response({"error": "Course not found."}, status=status.HTTP_404_NOT_FOUND)

        date_str = request.data.get("date")
        records  = request.data.get("records", [])

        if not date_str:
            return Response({"error": "date is required."}, status=status.HTTP_400_BAD_REQUEST)

        saved = 0
        alerts_sent = 0

        for rec in records:
            try:
                student = StudentProfile.objects.select_related("user").get(
                    user_id=rec["student_profile_id"]
                )

                # ── Snapshot absence count BEFORE this save ───────────────
                absences_before = _get_student_absence_count(student, course)
                status_before   = _classify_status(absences_before, course.max_absences)

                # ── Save / update the record ──────────────────────────────
                AttendanceRecord.objects.update_or_create(
                    course=course,
                    student=student,
                    date=date_str,
                    defaults={"status": rec["status"]},
                )
                saved += 1

                # ── Recount absences AFTER the save ───────────────────────
                absences_after = _get_student_absence_count(student, course)
                status_after   = _classify_status(absences_after, course.max_absences)

                # ── Notify student of any new absence (even first one) ────
                if rec.get("status") == "ABSENT" and absences_after == 1 and absences_before == 0:
                    _push_notification(
                        user=student.user,
                        notif_type=NotificationType.ABSENCE_INFO,
                        title=f"Absence recorded in {course.title}",
                        message=f"An absence has been recorded for {course.title}. You have {absences_after} of {course.max_absences} allowed.",
                        link="/student/attendance",
                        metadata={"course_id": course.id},
                    )

                # ── Handle threshold crossings (email + notif + deactivation) ─
                prev_status = _classify_status(absences_before, course.max_absences)
                new_status  = _classify_status(absences_after,  course.max_absences)
                thresholds  = ["OK", "WARNING", "DANGER"]
                if thresholds.index(new_status) > thresholds.index(prev_status):
                    _handle_absence_thresholds(student, course, absences_before, absences_after)
                    alerts_sent += 1

            except (StudentProfile.DoesNotExist, KeyError):
                continue

        return Response({"saved": saved, "alerts_sent": alerts_sent})


class SendDangerAlertsAPIView(APIView):
    """
    POST /api/teacher/courses/<course_id>/send-alerts/
    Manually re-sends email alerts to ALL WARNING and DANGER students
    in the given course. Useful as a one-click reminder from DangerZonePage.
    """
    permission_classes = [IsAuthenticated, IsAdminOrTeacher]

    def post(self, request, course_id):
        try:
            course = Course.objects.get(pk=course_id)
        except Course.DoesNotExist:
            return Response({"error": "Course not found."}, status=status.HTTP_404_NOT_FOUND)

        # Get all students enrolled in this course's filiere(s)
        filiere_ids = FiliereCourse.objects.filter(course=course).values_list("filiere_id", flat=True)
        students = (
            StudentProfile.objects
            .filter(filiere_id__in=filiere_ids)
            .select_related("user")
        )

        sent = 0
        skipped = 0
        for student in students:
            absences = _get_student_absence_count(student, course)
            alert_status = _classify_status(absences, course.max_absences)

            if alert_status in ("WARNING", "DANGER"):
                _send_absence_alert(
                    student.user, course,
                    absences, course.max_absences,
                    alert_status,
                )
                sent += 1
            else:
                skipped += 1

        return Response({
            "sent": sent,
            "skipped_ok": skipped,
            "message": f"Sent alerts to {sent} student{'s' if sent != 1 else ''}.",
        })



class AttendanceReportAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, course_id):
        try:
            course = Course.objects.select_related("teacher__user").get(pk=course_id)
        except Course.DoesNotExist:
            return Response({"error": "Course not found."}, status=status.HTTP_404_NOT_FOUND)

        user = request.user
        is_owner = (user.role == "TEACHER" and
                    hasattr(user, "teacher_profile") and
                    course.teacher == user.teacher_profile)
        if not (is_owner or user.role == "ADMIN"):
            return Response({"error": "Forbidden."}, status=status.HTTP_403_FORBIDDEN)

        # Optional filters — ?seance_id=N or ?date=YYYY-MM-DD
        date_filter   = request.query_params.get("date")
        seance_id_str = request.query_params.get("seance_id")

        records_qs = AttendanceRecord.objects.filter(course=course).select_related("student__user")
        if seance_id_str:
            records_qs = records_qs.filter(seance_id=seance_id_str)
        elif date_filter:
            records_qs = records_qs.filter(date=date_filter)

        records = records_qs.order_by("student__student_id", "date")

        from collections import defaultdict
        stats = defaultdict(lambda: {"present": 0, "absent": 0, "late": 0, "dates": []})
        for r in records:
            sid = r.student.student_id
            stats[sid]["name"] = (
                f"{r.student.user.last_name} {r.student.user.first_name}".strip()
                or r.student.user.username
            )
            stats[sid]["student_id"] = sid
            stats[sid][r.status.lower()] += 1
            stats[sid]["dates"].append((str(r.date), r.status))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Report"
        HEADER_BG = "7C3AED"
        ABSENT_BG = "7F1D1D"
        WARN_BG   = "78350F"
        WHITE     = "FFFFFF"

        def cell_style(cell, bold=False, bg=None, fg=WHITE, center=False, border=False):
            cell.font = Font(bold=bold, color=fg, size=11)
            if bg:
                cell.fill = PatternFill("solid", fgColor=bg)
            if center:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if border:
                thin = Side(style="thin", color="374151")
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.merge_cells("A1:G1")
        tc = ws["A1"]
        suffix = f"  [Séance #{seance_id_str}]" if seance_id_str else (f"  [{date_filter}]" if date_filter else "")
        tc.value = f"Attendance Report — {course.title}" + suffix
        cell_style(tc, bold=True, bg=HEADER_BG, center=True)
        tc.font = Font(bold=True, color=WHITE, size=14)
        ws.row_dimensions[1].height = 32

        ws.merge_cells("A2:G2")
        sc = ws["A2"]
        sc.value = (
            f"Teacher: {course.teacher.user.get_full_name() or course.teacher.user.username}"
            f"   |   Max absences: {course.max_absences}"
        )
        cell_style(sc, bg="374151", center=True)
        ws.row_dimensions[2].height = 22

        headers = ["#", "Student ID", "Full Name", "Present", "Absent", "Late", "Attendance %", "Status"]
        ws.append([])
        ws.append(headers)
        for col_idx, h in enumerate(headers, start=1):
            c = ws.cell(row=4, column=col_idx)
            c.value = h
            cell_style(c, bold=True, bg="374151", center=True, border=True)
            c.font = Font(bold=True, color=WHITE, size=11)
        ws.row_dimensions[4].height = 20

        for row_num, (sid, s) in enumerate(sorted(stats.items()), start=1):
            present = s["present"]
            absent  = s["absent"]
            late    = s["late"]
            total   = present + absent + late
            pct     = round((present / total * 100) if total else 0, 1)
            if absent >= course.max_absences:
                danger_status = "DANGER"
            elif absent >= course.max_absences - 1:
                danger_status = "WARNING"
            else:
                danger_status = "OK"
            row_data = [row_num, sid, s.get("name", ""), present, absent, late, f"{pct}%", danger_status]
            ws.append(row_data)
            excel_row = row_num + 4
            row_bg = ABSENT_BG if danger_status == "DANGER" else (WARN_BG if danger_status == "WARNING" else None)
            for col_idx in range(1, len(row_data) + 1):
                c = ws.cell(row=excel_row, column=col_idx)
                cell_style(c, bg=row_bg, center=(col_idx != 3), border=True)
                if not row_bg:
                    c.font = Font(color="D1D5DB", size=10)

        col_widths = [5, 14, 28, 10, 10, 10, 15, 12]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        safe_title = course.title.replace(' ', '_')
        if seance_id_str:
            filename = f"attendance_{safe_title}_seance_{seance_id_str}.xlsx"
        elif date_filter:
            filename = f"attendance_{safe_title}_{date_filter}.xlsx"
        else:
            filename = f"attendance_{safe_title}.xlsx"
        response = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class CourseAttendanceSummaryAPIView(APIView):
    """
    GET /api/teacher/courses/<course_id>/attendance-summary/
    Returns séances list with per-séance stats + per-student absence totals.
    """
    permission_classes = [IsAuthenticated, IsTeacherUserRole]

    def get(self, request, course_id):
        try:
            course = Course.objects.select_related("teacher").get(pk=course_id)
        except Course.DoesNotExist:
            return Response({"error": "Course not found."}, status=status.HTTP_404_NOT_FOUND)

        if course.teacher_id != request.user.pk:
            return Response({"error": "Forbidden."}, status=status.HTTP_403_FORBIDDEN)

        # ── Séances with inline stats ─────────────────────────────────────────
        seances = (
            Seance.objects
            .filter(course=course)
            .prefetch_related("attendance_records")
            .order_by("-date", "-start_time")
        )
        seances_data = []
        for s in seances:
            recs    = list(s.attendance_records.all())
            present = sum(1 for r in recs if r.status == "PRESENT")
            absent  = sum(1 for r in recs if r.status == "ABSENT")
            late    = sum(1 for r in recs if r.status == "LATE")
            total   = present + absent + late
            d       = _seance_to_dict(s, include_code=True)
            d.update({
                "present_count": present,
                "absent_count":  absent,
                "late_count":    late,
                "total_records": total,
                "rate": round((present / total) * 100) if total else 0,
            })
            seances_data.append(d)

        # ── Per-student absence summary ───────────────────────────────────────
        from collections import defaultdict
        records_qs = (
            AttendanceRecord.objects
            .filter(course=course)
            .select_related("student__user")
        )
        student_stats = defaultdict(lambda: {"present": 0, "absent": 0, "late": 0, "student_id": "", "full_name": ""})
        for rec in records_qs:
            sid = rec.student.student_id
            st  = student_stats[sid]
            st["student_id"] = sid
            st["full_name"]  = (
                f"{rec.student.user.first_name} {rec.student.user.last_name}".strip()
                or rec.student.user.username
            )
            st[rec.status.lower()] += 1

        students_list = sorted(student_stats.values(), key=lambda x: x["absent"], reverse=True)
        for st in students_list:
            total = st["present"] + st["absent"] + st["late"]
            st["total"]  = total
            st["rate"]   = round((st["present"] / total) * 100) if total else 0
            st["danger"] = st["absent"] >= course.max_absences

        return Response({
            "course_id":    course.id,
            "course_title": course.title,
            "max_absences": course.max_absences,
            "seances":      seances_data,
            "students":     students_list,
        })


def _push_notification(user, notif_type, title, message, link="", metadata=None):
    try:
        Notification.objects.create(
            user=user, type=notif_type, title=title,
            message=message, link=link, metadata=metadata or {},
        )
    except Exception as exc:
        logger.error(f"[Notif] Failed to create notification: {exc}")


class NotificationListAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        notifs = (
            Notification.objects
            .filter(user=request.user)
            .order_by("is_read", "-created_at")[:50]
        )
        data = [
            {
                "id": n.id, "type": n.type, "title": n.title,
                "message": n.message, "is_read": n.is_read,
                "link": n.link, "metadata": n.metadata,
                "created_at": n.created_at.isoformat(),
            }
            for n in notifs
        ]
        return Response({"notifications": data, "unread_count": sum(1 for n in notifs if not n.is_read)})


class NotificationReadAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, notif_id):
        try:
            notif = Notification.objects.get(pk=notif_id, user=request.user)
            notif.is_read = True
            notif.save(update_fields=["is_read"])
            return Response({"ok": True})
        except Notification.DoesNotExist:
            return Response({"error": "Not found."}, status=status.HTTP_404_NOT_FOUND)


class NotificationReadAllAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        updated = Notification.objects.filter(user=request.user, is_read=False).update(is_read=True)
        return Response({"marked_read": updated})


class NotificationUnreadCountAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        count = Notification.objects.filter(user=request.user, is_read=False).count()
        return Response({"unread_count": count})


class PlatformAssistantAPIView(APIView):
    """
    POST /api/platform-assistant/
    Body: { question: str, history: [{role, content}, ...] }
    General-purpose CampusEye platform assistant powered by Groq.
    Answers questions about how to use the platform, features, roles, etc.
    Available to all authenticated users (any role).
    """
    permission_classes = [IsAuthenticated]

    GROQ_API_URL = "https://api.groq.com/openai/v1/chat/completions"
    MODEL        = "llama-3.3-70b-versatile"

    SYSTEM_PROMPT = """You are the CampusEye platform assistant — a helpful, concise guide embedded in a smart academic attendance management system.

CampusEye has three roles:
- ADMIN: manages users, departments, filieres, courses; views platform stats.
- TEACHER: manages courses, uploads materials, scans attendance via face recognition, views danger zone (at-risk students), sends email alerts, marks attendance manually.
- STUDENT: views own attendance, downloads course materials, chats with the AI tutor, checks absence status.

Key features:
- Face-recognition attendance scanning (teacher scans a photo of the classroom)
- Manual attendance page (teacher marks each student present/absent/late)
- Danger Zone: students who reach WARNING or DANGER absence thresholds; teacher can send email alerts
- Course materials: teachers upload files (PDF, Word, etc.); students can view and download them
- AI Tutor chat (students): multi-agent system covering explanations, quizzes, exercises, and debugging help
- Attendance reports: downloadable Excel reports per course
- Notifications: real-time bell icon shows absence alerts, new materials, course assignments

Answer questions about CampusEye clearly and briefly. If the user asks something unrelated to the platform, politely redirect them."""

    def post(self, request):
        question = (request.data.get("question") or "").strip()
        history  = request.data.get("history") or []

        if not question:
            return Response({"answer": "Please type a question."}, status=status.HTTP_400_BAD_REQUEST)

        api_key = django_settings.GROQ_API_KEY if hasattr(django_settings, "GROQ_API_KEY") else os.environ.get("GROQ_API_KEY", "")
        if not api_key:
            return Response(
                {"answer": "The assistant is not configured yet. Please set GROQ_API_KEY in your .env file."},
                status=status.HTTP_200_OK,
            )

        # Build message list: system + trimmed history + current question
        messages = [{"role": "system", "content": self.SYSTEM_PROMPT}]
        for msg in history[-8:]:  # keep last 8 turns for context
            if isinstance(msg, dict) and msg.get("role") in ("user", "assistant") and msg.get("content"):
                messages.append({"role": msg["role"], "content": str(msg["content"])[:800]})
        messages.append({"role": "user", "content": question})

        payload = {
            "model":       self.MODEL,
            "messages":    messages,
            "max_tokens":  512,
            "temperature": 0.4,
        }
        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type":  "application/json",
        }

        try:
            resp = http_requests.post(
                self.GROQ_API_URL,
                json=payload,
                headers=headers,
                timeout=30,
            )
            resp.raise_for_status()
            answer = resp.json()["choices"][0]["message"]["content"]
            return Response({"answer": answer}, status=status.HTTP_200_OK)
        except http_requests.exceptions.Timeout:
            return Response(
                {"answer": "The AI took too long to respond. Please try again."},
                status=status.HTTP_200_OK,
            )
        except Exception as exc:
            logger.error(f"[PlatformAssistant] Groq error: {exc}")
            return Response(
                {"answer": "Something went wrong with the AI service. Please try again shortly."},
                status=status.HTTP_200_OK,
            )


# =============================================================================
# NOVAA AGENT EXECUTE ENDPOINT
# =============================================================================

class AgentExecuteAPIView(APIView):
    """
    POST /api/agent/execute/
    ────────────────────────
    Called by NOVAA (or any external AI) to execute CampusEye tasks via
    natural language. Supports a two-phase flow:

    Phase 1 — Natural language instruction:
      Body: { "instruction": "Create a student named Ahmed...", "params": {} }
      Response: { "status": "needs_info", "questions": [...] }
                 OR { "status": "executed", "result": {...} }

    Phase 2 — NOVAA supplies the missing parameters:
      Body: { "instruction": "...", "params": { "_tool": "create_user", "first_name": "Ahmed", ... } }
      Response: { "status": "executed", "result": {...} }

    The endpoint is accessible to ADMIN and TEACHER roles only.
    NOVAA must authenticate with a valid JWT token for the acting user.
    """
    permission_classes = [IsAuthenticated, IsAdminOrTeacher]

    def post(self, request):
        from .services.campuseye_agent_service import process_agent_request

        instruction = (request.data.get("instruction") or "").strip()
        params      = request.data.get("params") or {}

        if not instruction and not params.get("_tool"):
            return Response(
                {"status": "error", "message": "Provide an 'instruction' or a '_tool' in params."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        role   = request.user.role
        result = process_agent_request(
            instruction      = instruction,
            role             = role,
            params_override  = dict(params),   # copy so we don't mutate request.data
            requesting_user  = request.user,
        )

        http_status = status.HTTP_200_OK
        if result.get("status") == "error":
            http_status = status.HTTP_400_BAD_REQUEST

        return Response(result, status=http_status)


# =============================================================================
# SÉANCE SYSTEM
# =============================================================================

def _seance_to_dict(seance, include_stats=False, include_code=False):
    """
    Serialize a Seance to a plain dict.
    include_code: ONLY pass True for teacher/admin-facing responses — the
    check-in code must never reach students.
    """
    d = {
        "id":               seance.id,
        "course_id":        seance.course_id,
        "course_title":     seance.course.title,
        "date":             str(seance.date),
        "start_time":       seance.start_time.strftime("%H:%M"),
        "duration_minutes": seance.duration_minutes,
        "session_type":     seance.session_type,
        "tp_group":         seance.tp_group,
        "status":           seance.status,
        "notes":            seance.notes,
        "created_at":       seance.created_at.isoformat(),
        # Safe for everyone: whether a code is needed (NOT the code itself).
        "requires_code":    bool(seance.check_in_code),
    }
    if include_code:
        d["check_in_code"]      = seance.check_in_code
    if include_stats:
        records = seance.attendance_records.all()
        d["present_count"] = records.filter(status="PRESENT").count()
        d["absent_count"]  = records.filter(status="ABSENT").count()
        d["late_count"]    = records.filter(status="LATE").count()
        d["total_records"] = records.count()
    return d


def _get_seance_eligible_students(seance):
    """
    Return the StudentProfile queryset that should attend this séance.
    For COURS → all students enrolled in the course.
    For TP with GROUP_A/GROUP_B → filtered by tp_group.
    """
    filiere_ids = FiliereCourse.objects.filter(course=seance.course).values_list("filiere_id", flat=True)
    qs = StudentProfile.objects.filter(filiere_id__in=filiere_ids).select_related("user")
    if seance.session_type == SessionType.TP and seance.tp_group != TPGroup.NONE:
        qs = qs.filter(tp_group=seance.tp_group)
    return qs.order_by("user__last_name", "user__first_name")


def _auto_activate_scheduled_seance(seance):
    """If a SCHEDULED seance is within 5 minutes of its start time, activate it."""
    if seance.status != SeanceStatus.SCHEDULED:
        return False
    now = datetime.now()
    seance_start = datetime.combine(seance.date, seance.start_time)
    if now >= seance_start - timedelta(minutes=5):
        seance.status = SeanceStatus.ACTIVE
        seance.save(update_fields=["status"])
        return True
    return False


def _auto_complete_expired_seance(seance):
    """
    If the seance is ACTIVE and its end time has passed, mark it COMPLETED
    and auto-create ABSENT records for students who never checked in.
    Returns True if the seance was completed by this call.
    """
    if seance.status != SeanceStatus.ACTIVE:
        return False
    now = datetime.now()
    seance_end = datetime.combine(seance.date, seance.start_time) + timedelta(minutes=seance.duration_minutes)
    if now < seance_end:
        return False

    seance.status = SeanceStatus.COMPLETED
    seance.save(update_fields=["status"])

    eligible   = _get_seance_eligible_students(seance)
    already_in = set(AttendanceRecord.objects.filter(seance=seance).values_list("student_id", flat=True))
    for student in eligible:
        if student.pk not in already_in:
            absences_before = _get_student_absence_count(student, seance.course)
            AttendanceRecord.objects.create(
                course=seance.course,
                student=student,
                seance=seance,
                date=seance.date,
                status="ABSENT",
            )
            absences_after = _get_student_absence_count(student, seance.course)
            _handle_absence_thresholds(student, seance.course, absences_before, absences_after)
    return True


class SeanceListCreateAPIView(APIView):
    """
    GET  /api/teacher/courses/<course_id>/seances/   — list séances for a course
    POST /api/teacher/courses/<course_id>/seances/   — create one (or two for TP both groups)
    """
    permission_classes = [IsAuthenticated, IsTeacherUserRole]

    def _get_course(self, request, course_id):
        try:
            course = Course.objects.select_related("teacher").get(pk=course_id)
        except Course.DoesNotExist:
            return None, Response({"error": "Course not found."}, status=status.HTTP_404_NOT_FOUND)
        if course.teacher_id != request.user.pk:
            return None, Response({"error": "Forbidden."}, status=status.HTTP_403_FORBIDDEN)
        return course, None

    def get(self, request, course_id):
        course, err = self._get_course(request, course_id)
        if err:
            return err

        seances = list(Seance.objects.filter(course=course).prefetch_related("attendance_records").select_related("course"))
        for s in seances:
            _auto_activate_scheduled_seance(s)
            _auto_complete_expired_seance(s)
        status_filter = request.query_params.get("status")
        if status_filter:
            seances = [s for s in seances if s.status == status_filter.upper()]

        return Response([_seance_to_dict(s, include_stats=True, include_code=True) for s in seances])

    def post(self, request, course_id):
        course, err = self._get_course(request, course_id)
        if err:
            return err

        data            = request.data
        date_str        = data.get("date")
        start_time_str  = data.get("start_time")
        duration        = data.get("duration_minutes", 60)
        session_type    = data.get("session_type", SessionType.COURS)
        tp_group_val    = data.get("tp_group", TPGroup.NONE)
        notes           = data.get("notes", "")
        # Check-in code: teacher may supply one, else auto-generate. Required for
        # the student to unlock face recognition; never sent to students.
        check_in_code   = (data.get("check_in_code") or "").strip().upper() or generate_seance_code()

        if not date_str or not start_time_str:
            return Response({"error": "date and start_time are required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            seance_date = datetime.strptime(date_str, "%Y-%m-%d").date()
            start_time  = datetime.strptime(start_time_str, "%H:%M").time()
            duration    = int(duration)
        except ValueError:
            return Response({"error": "Invalid date/time format."}, status=status.HTTP_400_BAD_REQUEST)

        created_seances = []

        # If TP and teacher selected "both groups" → create two back-to-back séances
        if session_type == SessionType.TP and tp_group_val == "BOTH":
            groups = [TPGroup.GROUP_A, TPGroup.GROUP_B]
            current_start = datetime.combine(seance_date, start_time)
            for grp in groups:
                s = Seance.objects.create(
                    course=course,
                    date=seance_date,
                    start_time=current_start.time(),
                    duration_minutes=duration,
                    session_type=SessionType.TP,
                    tp_group=grp,
                    notes=notes,
                    check_in_code=check_in_code,
                    created_by=request.user,
                )
                created_seances.append(s)
                # Next group starts immediately after this one ends
                current_start += timedelta(minutes=duration)
        else:
            # Validate tp_group for TP
            if session_type == SessionType.TP and tp_group_val == TPGroup.NONE:
                tp_group_val = TPGroup.GROUP_A  # default fallback

            s = Seance.objects.create(
                course=course,
                date=seance_date,
                start_time=start_time,
                duration_minutes=duration,
                session_type=session_type,
                tp_group=tp_group_val if session_type == SessionType.TP else TPGroup.NONE,
                notes=notes,
                check_in_code=check_in_code,
                created_by=request.user,
            )
            created_seances.append(s)

        # Auto-activate seances scheduled to start within the next 5 minutes
        now = datetime.now()
        for s in created_seances:
            seance_start = datetime.combine(s.date, s.start_time)
            if now >= seance_start - timedelta(minutes=5):
                s.status = SeanceStatus.ACTIVE
                s.save(update_fields=["status"])

        # Notify eligible students for each created séance
        for s in created_seances:
            eligible = _get_seance_eligible_students(s)
            grp_label = f" — {s.tp_group.replace('_', ' ')}" if s.tp_group != TPGroup.NONE else ""
            type_label = "Cours" if s.session_type == SessionType.COURS else f"TP{grp_label}"
            for student in eligible:
                _push_notification(
                    user=student.user,
                    notif_type=NotificationType.SEANCE_CREATED,
                    title=f"📅 Nouvelle séance — {course.title}",
                    message=f"{type_label} prévu le {s.date} à {s.start_time.strftime('%H:%M')} ({s.duration_minutes} min).",
                    link="/student/seances",
                    metadata={"seance_id": s.id, "course_id": course.id},
                )

        return Response(
            [_seance_to_dict(s, include_code=True) for s in created_seances],
            status=status.HTTP_201_CREATED,
        )


class SeanceDetailAPIView(APIView):
    """
    GET    /api/teacher/seances/<seance_id>/   — detail + roster
    PATCH  /api/teacher/seances/<seance_id>/   — update notes / date / time
    DELETE /api/teacher/seances/<seance_id>/   — delete (only if SCHEDULED)
    """
    permission_classes = [IsAuthenticated, IsTeacherUserRole]

    def _get_seance(self, request, seance_id):
        try:
            s = Seance.objects.select_related("course__teacher").get(pk=seance_id)
        except Seance.DoesNotExist:
            return None, Response({"error": "Séance not found."}, status=status.HTTP_404_NOT_FOUND)
        if int(s.course.teacher_id) != int(request.user.pk) and getattr(request.user, "role", None) != "ADMIN":
            return None, Response({"error": "Forbidden."}, status=status.HTTP_403_FORBIDDEN)
        return s, None

    def _build_roster(self, seance):
        eligible = _get_seance_eligible_students(seance)
        records = {
            r.student_id: r
            for r in AttendanceRecord.objects.filter(seance=seance).select_related("student")
        }
        roster = []
        for student in eligible:
            rec = records.get(student.pk)
            roster.append({
                "student_id": student.student_id,
                "user_id":    student.pk,
                "full_name":  f"{student.user.first_name} {student.user.last_name}".strip() or student.user.username,
                "tp_group":   student.tp_group,
                "status":     rec.status if rec else "ABSENT",
                "record_id":  rec.id if rec else None,
            })
        return roster

    def get(self, request, seance_id):
        seance, err = self._get_seance(request, seance_id)
        if err:
            return err
        _auto_activate_scheduled_seance(seance)
        _auto_complete_expired_seance(seance)
        data = _seance_to_dict(seance, include_stats=True, include_code=True)
        data["roster"] = self._build_roster(seance)
        return Response(data)

    def patch(self, request, seance_id):
        seance, err = self._get_seance(request, seance_id)
        if err:
            return err
        if seance.status == SeanceStatus.COMPLETED:
            return Response({"error": "Cannot edit a completed séance."}, status=status.HTTP_400_BAD_REQUEST)

        allowed = ["date", "start_time", "duration_minutes", "notes", "session_type", "tp_group"]
        for field in allowed:
            if field in request.data:
                setattr(seance, field, request.data[field])
        seance.save()
        # Refresh so time/date fields are proper Python objects (not raw strings)
        # before _seance_to_dict calls .strftime() on them
        seance.refresh_from_db()
        return Response(_seance_to_dict(seance, include_code=True))

    def delete(self, request, seance_id):
        seance, err = self._get_seance(request, seance_id)
        if err:
            return err
        if seance.status == SeanceStatus.ACTIVE:
            return Response({"error": "Cannot delete an active séance."}, status=status.HTTP_400_BAD_REQUEST)
        seance.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class SeanceStartAPIView(APIView):
    """POST /api/teacher/seances/<seance_id>/start/  — activate a séance"""
    permission_classes = [IsAuthenticated, IsTeacherUserRole]

    def post(self, request, seance_id):
        try:
            seance = Seance.objects.select_related("course__teacher").get(pk=seance_id)
        except Seance.DoesNotExist:
            return Response({"error": "Séance not found."}, status=status.HTTP_404_NOT_FOUND)
        if seance.course.teacher_id != request.user.pk:
            return Response({"error": "Forbidden."}, status=status.HTTP_403_FORBIDDEN)
        if seance.status != SeanceStatus.SCHEDULED:
            return Response({"error": f"Séance is already {seance.status}."}, status=status.HTTP_400_BAD_REQUEST)

        seance.status = SeanceStatus.ACTIVE
        seance.save(update_fields=["status"])

        # Notify eligible students that the séance has started (code is NOT included)
        eligible = _get_seance_eligible_students(seance)
        for student in eligible:
            _push_notification(
                user=student.user,
                notif_type=NotificationType.SEANCE_STARTED,
                title=f"🟢 Séance démarrée — {seance.course.title}",
                message=f"La séance a commencé. Pointez votre présence maintenant !",
                link="/student/seances",
                metadata={"seance_id": seance.id},
            )

        return Response(_seance_to_dict(seance, include_code=True))


class SeanceEndAPIView(APIView):
    """
    POST /api/teacher/seances/<seance_id>/end/
    Completes the séance and auto-creates ABSENT records for students who never checked in.
    """
    permission_classes = [IsAuthenticated, IsTeacherUserRole]

    def post(self, request, seance_id):
        try:
            seance = Seance.objects.select_related("course__teacher", "course").get(pk=seance_id)
        except Seance.DoesNotExist:
            return Response({"error": "Séance not found."}, status=status.HTTP_404_NOT_FOUND)
        if seance.course.teacher_id != request.user.pk:
            return Response({"error": "Forbidden."}, status=status.HTTP_403_FORBIDDEN)
        if seance.status != SeanceStatus.ACTIVE:
            return Response({"error": "Séance is not active."}, status=status.HTTP_400_BAD_REQUEST)

        seance.status = SeanceStatus.COMPLETED
        seance.save(update_fields=["status"])

        # Auto-mark absent anyone who never checked in
        eligible       = _get_seance_eligible_students(seance)
        already_in     = set(AttendanceRecord.objects.filter(seance=seance).values_list("student_id", flat=True))
        absent_created = 0
        for student in eligible:
            if student.pk not in already_in:
                absences_before = _get_student_absence_count(student, seance.course)
                AttendanceRecord.objects.create(
                    course=seance.course,
                    student=student,
                    seance=seance,
                    date=seance.date,
                    status="ABSENT",
                )
                absent_created += 1
                absences_after = _get_student_absence_count(student, seance.course)
                _handle_absence_thresholds(student, seance.course, absences_before, absences_after)

        return Response({
            **_seance_to_dict(seance, include_stats=True, include_code=True),
            "auto_absent_created": absent_created,
        })


class SeanceManualAttendanceAPIView(APIView):
    """
    POST /api/teacher/seances/<seance_id>/manual/
    Body: { records: [{ student_id: "...", status: "PRESENT|ABSENT|LATE" }, ...] }
    Teacher manually sets attendance within a séance (creates or updates per-student records).
    """
    permission_classes = [IsAuthenticated, IsTeacherUserRole]

    def post(self, request, seance_id):
        try:
            seance = Seance.objects.select_related("course__teacher", "course").get(pk=seance_id)
        except Seance.DoesNotExist:
            return Response({"error": "Séance not found."}, status=status.HTTP_404_NOT_FOUND)
        if seance.course.teacher_id != request.user.pk:
            return Response({"error": "Forbidden."}, status=status.HTTP_403_FORBIDDEN)
        if seance.status == SeanceStatus.CANCELLED:
            return Response({"error": "Séance is cancelled."}, status=status.HTTP_400_BAD_REQUEST)

        _auto_complete_expired_seance(seance)

        records_data = request.data.get("records", [])
        saved = 0
        for entry in records_data:
            sid        = entry.get("student_id")
            status_val = entry.get("status", "PRESENT").upper()
            if status_val not in ("PRESENT", "ABSENT", "LATE"):
                continue
            try:
                student = StudentProfile.objects.select_related("user").get(student_id=sid)
            except StudentProfile.DoesNotExist:
                continue

            absences_before = _get_student_absence_count(student, seance.course)

            AttendanceRecord.objects.update_or_create(
                seance=seance,
                student=student,
                defaults={
                    "course": seance.course,
                    "date":   seance.date,
                    "status": status_val,
                },
            )
            saved += 1

            absences_after = _get_student_absence_count(student, seance.course)
            _handle_absence_thresholds(student, seance.course, absences_before, absences_after)

        # Return updated roster
        roster = SeanceDetailAPIView()._build_roster(seance)
        return Response({"saved": saved, "roster": roster})


class SeanceFaceScanAPIView(APIView):
    """
    POST /api/teacher/seances/<seance_id>/scan/
    Runs face recognition on an uploaded image, marks recognized students PRESENT
    for the given séance, and returns the updated roster.
    """
    permission_classes = [IsAuthenticated, IsTeacherUserRole]

    def post(self, request, seance_id):
        try:
            seance = Seance.objects.select_related("course__teacher", "course").get(pk=seance_id)
        except Seance.DoesNotExist:
            return Response({"error": "Séance not found."}, status=status.HTTP_404_NOT_FOUND)
        if seance.course.teacher_id != request.user.pk:
            return Response({"error": "Forbidden."}, status=status.HTTP_403_FORBIDDEN)
        if seance.status not in (SeanceStatus.ACTIVE, SeanceStatus.SCHEDULED):
            return Response({"error": "Séance is not active."}, status=status.HTTP_400_BAD_REQUEST)

        image = request.FILES.get("image")
        if not image:
            return Response({"error": "image is required."}, status=status.HTTP_400_BAD_REQUEST)

        # Auto-activate if still scheduled
        if seance.status == SeanceStatus.SCHEDULED:
            seance.status = SeanceStatus.ACTIVE
            seance.save(update_fields=["status"])

        # Run face recognition (reuse existing service)
        result = recognize_and_mark_attendance(image, seance.course_id)

        recognized_student_ids = {s["student_id"] for s in result.get("recognized_students", [])}

        # For each recognized student, create/update attendance record tied to the séance
        eligible   = _get_seance_eligible_students(seance)
        eligible_ids = {s.student_id: s for s in eligible}
        newly_marked = 0

        for sid in recognized_student_ids:
            student = eligible_ids.get(sid)
            if not student:
                continue  # recognized but not in this séance's group — skip
            _, created = AttendanceRecord.objects.update_or_create(
                seance=seance,
                student=student,
                defaults={
                    "course": seance.course,
                    "date":   seance.date,
                    "status": "PRESENT",
                },
            )
            if created:
                newly_marked += 1

        # Build roster
        existing_records = {
            r.student_id: r.status
            for r in AttendanceRecord.objects.filter(seance=seance)
        }
        roster = []
        for student in eligible:
            rec_status = existing_records.get(student.pk, "ABSENT")
            roster.append({
                "student_id":      student.student_id,
                "full_name":       f"{student.user.first_name} {student.user.last_name}".strip() or student.user.username,
                "status":          rec_status,
                "just_recognized": student.student_id in recognized_student_ids,
            })

        return Response({
            "success":            True,
            "faces_detected":     result.get("faces_detected", 0),
            "newly_marked":       newly_marked,
            "recognized_in_group": len([s for s in recognized_student_ids if s in eligible_ids]),
            "roster":             roster,
            "total_enrolled":     len(roster),
            "present_count":      sum(1 for r in roster if r["status"] == "PRESENT"),
            "absent_count":       sum(1 for r in roster if r["status"] == "ABSENT"),
            "late_count":         sum(1 for r in roster if r["status"] == "LATE"),
            "seance":             _seance_to_dict(seance, include_code=True),
        })


# ── Student-side séance APIs ──────────────────────────────────────────────────

class StudentSeanceListAPIView(APIView):
    """
    GET /api/student/seances/
    Returns all séances for the authenticated student filtered by:
    - courses they are enrolled in (via filière)
    - their tp_group (they only see séances matching their group, or COURS séances)
    Optional: ?upcoming=true — only SCHEDULED/ACTIVE ones
    """
    permission_classes = [IsAuthenticated, IsStudentUserRole]

    def get(self, request):
        try:
            student = StudentProfile.objects.select_related("filiere").get(user=request.user)
        except StudentProfile.DoesNotExist:
            return Response({"error": "Student profile not found."}, status=status.HTTP_404_NOT_FOUND)

        # All courses the student is enrolled in via their filière
        course_ids = FiliereCourse.objects.filter(
            filiere=student.filiere
        ).values_list("course_id", flat=True)

        # Séances for those courses that are either:
        # - type COURS (tp_group NONE) → all students
        # - type TP matching the student's tp_group
        from django.db.models import Q
        seances = Seance.objects.filter(course_id__in=course_ids).filter(
            Q(tp_group=TPGroup.NONE) | Q(tp_group=student.tp_group)
        ).select_related("course").prefetch_related("attendance_records")

        upcoming = request.query_params.get("upcoming")
        if upcoming == "true":
            seances = seances.filter(status__in=[SeanceStatus.SCHEDULED, SeanceStatus.ACTIVE])

        result = []
        for s in seances:
            d = _seance_to_dict(s)
            # Check if this student already has an attendance record for this séance
            rec = AttendanceRecord.objects.filter(seance=s, student=student).first()
            d["my_status"] = rec.status if rec else None

            # Compute whether check-in is currently allowed
            # Window: 5 min before start → end of séance
            now = datetime.now()
            seance_start = datetime.combine(s.date, s.start_time)
            seance_end   = seance_start + timedelta(minutes=s.duration_minutes)
            window_open  = seance_start - timedelta(minutes=5)
            d["can_check_in"] = (
                s.status == SeanceStatus.ACTIVE and
                window_open <= now <= seance_end and
                rec is None
            )
            d["window_opens_at"] = window_open.isoformat()
            d["seance_ends_at"]  = seance_end.isoformat()
            result.append(d)

        return Response(result)


class SeanceVerifyCodeAPIView(APIView):
    """
    POST /api/student/seances/<seance_id>/verify-code/
    Body: { code }
    Validates the séance check-in code WITHOUT doing face recognition, so the
    frontend can unlock the camera only after the correct code is entered.
    Returns {"valid": bool}. The code itself is never returned.
    """
    permission_classes = [IsAuthenticated, IsStudentUserRole]

    def post(self, request, seance_id):
        try:
            seance = Seance.objects.get(pk=seance_id)
        except Seance.DoesNotExist:
            return Response({"error": "Séance not found."}, status=status.HTTP_404_NOT_FOUND)

        required_code = (seance.check_in_code or "").strip()
        # No code configured → nothing to verify, camera can open directly.
        if not required_code:
            return Response({"valid": True, "requires_code": False})

        submitted = (request.data.get("code") or "").strip().upper()
        if not submitted:
            return Response({"valid": False, "requires_code": True,
                             "error": "Veuillez entrer le code de la séance."},
                            status=status.HTTP_400_BAD_REQUEST)
        if submitted != required_code.upper():
            return Response({"valid": False, "requires_code": True,
                             "error": "Code de séance incorrect."},
                            status=status.HTTP_400_BAD_REQUEST)
        return Response({"valid": True, "requires_code": True})


class StudentCheckInAPIView(APIView):
    """
    POST /api/student/seances/<seance_id>/check-in/
    Student submits a selfie; we verify it matches only THEIR stored face encoding.
    On success → AttendanceRecord with PRESENT (or LATE if > 15 min past start).
    """
    permission_classes = [IsAuthenticated, IsStudentUserRole]

    def post(self, request, seance_id):
        import face_recognition as fr
        import numpy as np

        try:
            seance = Seance.objects.select_related("course").get(pk=seance_id)
        except Seance.DoesNotExist:
            return Response({"error": "Séance not found."}, status=status.HTTP_404_NOT_FOUND)

        try:
            student = StudentProfile.objects.get(user=request.user)
        except StudentProfile.DoesNotExist:
            return Response({"error": "Student profile not found."}, status=status.HTTP_404_NOT_FOUND)

        # Time window check
        now          = datetime.now()
        seance_start = datetime.combine(seance.date, seance.start_time)
        seance_end   = seance_start + timedelta(minutes=seance.duration_minutes)
        window_open  = seance_start - timedelta(minutes=5)

        # Face-match threshold. distance ∈ [0, ~1]; lower = closer match.
        THRESHOLD = 0.5
        # Map a distance to a 0–100 confidence %, clamped.
        def _confidence(dist):
            return max(0, min(100, round((1 - (dist / THRESHOLD)) * 100)))

        if seance.status not in (SeanceStatus.ACTIVE, SeanceStatus.SCHEDULED):
            return Response({"error": "Cette séance n'est pas active.", "reason": "NOT_ACTIVE", "matched": False}, status=status.HTTP_400_BAD_REQUEST)
        if now < window_open:
            return Response({"error": f"La séance n'est pas encore accessible. Revenez après {window_open.strftime('%H:%M')}.", "reason": "WINDOW_NOT_OPEN", "matched": False}, status=status.HTTP_400_BAD_REQUEST)
        if now > seance_end:
            return Response({"error": "La séance est terminée.", "reason": "WINDOW_CLOSED", "matched": False}, status=status.HTTP_400_BAD_REQUEST)

        # Already checked in?
        existing = AttendanceRecord.objects.filter(seance=seance, student=student).first()
        if existing:
            return Response({
                "error": "Vous avez déjà pointé votre présence pour cette séance.",
                "reason": "ALREADY_CHECKED_IN", "matched": False,
                "status": existing.status,
            }, status=status.HTTP_400_BAD_REQUEST)

        # Check group eligibility
        if seance.tp_group != TPGroup.NONE and student.tp_group != seance.tp_group:
            return Response({"error": "Vous n'êtes pas dans le groupe de cette séance.", "reason": "WRONG_GROUP", "matched": False}, status=status.HTTP_403_FORBIDDEN)

        # ── Séance code gate — must be correct BEFORE face recognition runs ───
        # The code is given by the teacher in class; it is never pushed to
        # students, so entering it proves physical presence in the room.
        required_code = (seance.check_in_code or "").strip()
        if required_code:
            submitted = (request.data.get("code") or "").strip().upper()
            if not submitted:
                return Response({"error": "Veuillez d'abord entrer le code de la séance.", "reason": "CODE_REQUIRED", "matched": False}, status=status.HTTP_400_BAD_REQUEST)
            if submitted != required_code.upper():
                return Response({"error": "Code de séance incorrect.", "reason": "CODE_INVALID", "matched": False}, status=status.HTTP_400_BAD_REQUEST)

        # Must have face encoding registered
        logger.info(
            "[CheckIn] student=%s seance=%s has_encoding=%s enc_len=%s",
            student.student_id, seance_id,
            bool(student.face_encoding),
            len(student.face_encoding) if student.face_encoding else 0,
        )
        if not student.face_encoding:
            return Response({"error": "Votre visage n'est pas encore enregistré. Contactez un administrateur.", "reason": "NO_REGISTERED_FACE", "matched": False}, status=status.HTTP_400_BAD_REQUEST)

        # Get uploaded selfie
        image_file = request.FILES.get("image")
        if not image_file:
            return Response({"error": "Aucune image fournie.", "reason": "NO_IMAGE", "matched": False}, status=status.HTTP_400_BAD_REQUEST)

        try:
            img_array   = fr.load_image_file(image_file)
            face_locs   = fr.face_locations(img_array)
            if not face_locs:
                return Response({"error": "Aucun visage détecté dans l'image.", "reason": "NO_FACE_DETECTED", "matched": False})
            if len(face_locs) > 1:
                return Response({"error": "Plusieurs visages détectés. Assurez-vous d'être seul dans le cadre.", "reason": "MULTIPLE_FACES", "matched": False})
            encodings   = fr.face_encodings(img_array, face_locs)
            if not encodings:
                return Response({"error": "Impossible d'analyser le visage. Améliorez l'éclairage et réessayez.", "reason": "ENCODING_FAILED", "matched": False})

            known_encoding  = np.array(student.face_encoding)
            selfie_encoding = encodings[0]
            distance = fr.face_distance([known_encoding], selfie_encoding)[0]
            confidence = _confidence(distance)

            logger.info(
                "[CheckIn] student=%s faces_found=%d distance=%.4f confidence=%d%% threshold=%.2f -> %s",
                student.student_id, len(face_locs), float(distance), confidence, THRESHOLD,
                "MATCH" if distance <= THRESHOLD else "REJECT",
            )

            if distance > THRESHOLD:
                return Response({
                    "error": "Visage non reconnu — ce n'est pas vous, ou la qualité est insuffisante.",
                    "reason": "NOT_RECOGNIZED", "matched": False,
                    "distance": float(distance), "confidence": confidence,
                })
        except Exception as exc:
            logger.error(f"[StudentCheckIn] Face error: {exc}")
            return Response({"error": "Erreur lors de la reconnaissance faciale.", "reason": "RECOGNITION_ERROR", "matched": False}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # Determine status: PRESENT or LATE (>15 min past start)
        minutes_late = (now - seance_start).total_seconds() / 60
        attendance_status = "LATE" if minutes_late > 15 else "PRESENT"

        AttendanceRecord.objects.create(
            course=seance.course,
            student=student,
            seance=seance,
            date=seance.date,
            status=attendance_status,
        )

        return Response({
            "matched":     True,
            "status":      attendance_status,
            "reason":      "SUCCESS",
            "message":     "Présence enregistrée !" if attendance_status == "PRESENT" else "Présence enregistrée (en retard).",
            "distance":    float(distance),
            "confidence":  confidence,
            "minutes_late": round(minutes_late) if attendance_status == "LATE" else 0,
            "checked_in_at": now.strftime("%H:%M"),
        })


# ══════════════════════════════════════════════════════════════════════════════
# NOVAA AI TUTOR — unified endpoint for all roles
# POST /api/ai/ask/
# ══════════════════════════════════════════════════════════════════════════════

class NovaaAskAPIView(APIView):
    """
    Unified AI tutor endpoint powered by NOVAA's upgraded brain.
    Replaces the old /api/chat/ask/ (student-only) with a multi-role endpoint.

    Request body:
      question      str  required
      course_id     int  optional
      session_id    int  optional
      mode          str  optional
      file_context  str  optional
      student_id    str  optional (TEACHER only: for email draft context)
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        question = (request.data.get("question") or "").strip()
        if not question:
            return Response(
                {"error": "question is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        user       = request.user
        role       = getattr(user, "role", "STUDENT")
        user_name  = user.get_full_name() or user.username
        course_id  = request.data.get("course_id")
        session_id = request.data.get("session_id")
        mode       = request.data.get("mode")
        file_context = request.data.get("file_context")

        # Build live platform context from DB
        platform_context = get_platform_context(
            user_id=user.id,
            role=role,
            course_id=course_id,
        )

        # Teachers: optionally attach per-student absence detail for email drafts
        student_id_for_email = request.data.get("student_id")
        if role == "TEACHER" and student_id_for_email:
            email_ctx = get_email_context_for_student(
                teacher_user_id=user.id,
                student_id_str=str(student_id_for_email),
            )
            if email_ctx:
                platform_context += "\n\n--- TARGET STUDENT FOR EMAIL ---\n" + email_ctx

        try:
            result = ask_novaa(
                question=question,
                user_id=user.id,
                role=role,
                user_name=user_name,
                course_id=course_id,
                session_id=session_id,
                mode=mode,
                file_context=file_context,
                platform_context=platform_context,
            )
        except Exception as exc:
            logger.error("[NovaaAskAPIView] ask_novaa failed for user %s: %s", user.id, exc)
            return Response(
                {"error": "The AI tutor encountered an unexpected error. Please try again."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        # Always return 200 — success/failure is communicated in result["success"].
        # Returning 4xx causes the frontend to show "Connection lost" instead of
        # the actual error message from the AI/action executor.
        return Response(result, status=status.HTTP_200_OK)


class NovaaAskStreamAPIView(APIView):
    """
    Streaming variant of NovaaAskAPIView — Server-Sent Events (SSE).
    POST /api/ai/ask/stream/  (same body as /api/ai/ask/)

    Emits a sequence of `data: {json}\\n\\n` events:
      {"type":"token","text":"..."}            incremental answer text
      {"type":"done", "answer":"...", ...}     final metadata (mode, sources, followups, verification)
      {"type":"error","answer":"..."}          failure
    The client appends tokens live, then finalises on `done`.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        question = (request.data.get("question") or "").strip()
        if not question:
            return Response({"error": "question is required."},
                            status=status.HTTP_400_BAD_REQUEST)

        user       = request.user
        role       = getattr(user, "role", "STUDENT")
        user_name  = user.get_full_name() or user.username
        course_id  = request.data.get("course_id")
        session_id = request.data.get("session_id")
        mode       = request.data.get("mode")
        file_context = request.data.get("file_context")

        platform_context = get_platform_context(
            user_id=user.id, role=role, course_id=course_id,
        )
        student_id_for_email = request.data.get("student_id")
        if role == "TEACHER" and student_id_for_email:
            email_ctx = get_email_context_for_student(
                teacher_user_id=user.id, student_id_str=str(student_id_for_email),
            )
            if email_ctx:
                platform_context += "\n\n--- TARGET STUDENT FOR EMAIL ---\n" + email_ctx

        def event_stream():
            try:
                for event in ask_novaa_stream(
                    question=question,
                    user_id=user.id,
                    role=role,
                    user_name=user_name,
                    course_id=course_id,
                    session_id=session_id,
                    mode=mode,
                    file_context=file_context,
                    platform_context=platform_context,
                ):
                    yield f"data: {json.dumps(event)}\n\n"
            except Exception as exc:
                logger.error("[NovaaAskStream] stream failed for user %s: %s", user.id, exc)
                err = {"type": "error",
                       "answer": "The AI tutor encountered an unexpected error. Please try again."}
                yield f"data: {json.dumps(err)}\n\n"

        response = StreamingHttpResponse(event_stream(), content_type="text/event-stream")
        response["Cache-Control"] = "no-cache"
        response["X-Accel-Buffering"] = "no"  # disable proxy buffering (nginx)
        return response


# ── HUD Stats endpoint (NOVAA Desktop integration) ────────────────────────────
class HudStatsAPIView(APIView):
    """
    Lightweight stats endpoint consumed by the NOVAA desktop HUD.
    Protected by a shared token in the X-HUD-Token header instead of JWT,
    so the desktop app can call it without a user login session.
    """
    permission_classes = []  # token-based auth handled manually

    def get(self, request):
        from django.conf import settings as django_settings

        # ── Token check ───────────────────────────────────────────────────
        expected = getattr(django_settings, "HUD_SECRET_TOKEN", "novaa-hud-2024")
        provided = request.META.get("HTTP_X_HUD_TOKEN", "")
        if provided != expected:
            return Response({"detail": "Forbidden"}, status=status.HTTP_403_FORBIDDEN)

        # ── Counts ────────────────────────────────────────────────────────
        total_students  = User.objects.filter(role="STUDENT").count()
        total_teachers  = User.objects.filter(role="TEACHER").count()
        total_courses   = Course.objects.count()

        # Active seances (started but not yet ended)
        active_sessions = Seance.objects.filter(status="ACTIVE").count()

        # Danger zone: students whose absence count >= course max_absences
        danger_zone_count = 0
        for sp in StudentProfile.objects.select_related("user", "filiere").prefetch_related(
            "filiere__filiere_courses__course"
        ):
            if sp.filiere is None:
                continue
            student = sp.user
            in_danger = False
            for fc in sp.filiere.filiere_courses.select_related("course").all():
                try:
                    course = fc.course
                    absences = AttendanceRecord.objects.filter(
                        student=student, course=course, status="ABSENT"
                    ).count()
                    if absences >= course.max_absences:
                        in_danger = True
                        break
                except Exception:
                    continue
            if in_danger:
                danger_zone_count += 1

        # 30-day attendance rate
        from django.utils.timezone import now as _now
        cutoff = _now() - timedelta(days=30)
        recent_records = AttendanceRecord.objects.filter(date__gte=cutoff)
        total_recent = recent_records.count()
        present_recent = recent_records.filter(status="PRESENT").count()
        attendance_rate_30d = round((present_recent / total_recent * 100), 1) if total_recent else 0.0

        # Pending face registration requests
        try:
            pending_face_requests = FaceRegistrationRequest.objects.filter(status="PENDING").count()
        except Exception:
            pending_face_requests = 0

        return Response({
            "total_students":        total_students,
            "total_teachers":        total_teachers,
            "total_courses":         total_courses,
            "active_sessions":       active_sessions,
            "danger_zone_count":     danger_zone_count,
            "attendance_rate_30d":   attendance_rate_30d,
            "pending_face_requests": pending_face_requests,
        }, status=status.HTTP_200_OK)

# ── NOVAA PDF Export ──────────────────────────────────────────────────────────
class NovaaGeneratePDFView(APIView):
    """
    POST /api/ai/pdf/
    Body: { content, title, doc_type, course_title }
    Returns a PDF file as an attachment.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        content      = (request.data.get("content")      or "").strip()
        title        = (request.data.get("title")        or "NOVAA Document").strip()
        doc_type     = (request.data.get("doc_type")     or "general").strip()
        course_title = (request.data.get("course_title") or "").strip()

        if not content:
            return Response({"error": "content is required."}, status=status.HTTP_400_BAD_REQUEST)

        user   = request.user
        role   = getattr(user, "role", "STUDENT")
        author = user.get_full_name() or user.username

        try:
            from attendance.services.novaa_pdf_service import generate_pdf
            pdf_bytes = generate_pdf(
                content=content,
                title=title,
                author=author,
                role=role,
                doc_type=doc_type,
                course_title=course_title,
            )
        except RuntimeError as exc:
            return Response({"error": str(exc)}, status=status.HTTP_503_SERVICE_UNAVAILABLE)
        except Exception as exc:
            logger.error("[NovaaGeneratePDF] failed: %s", exc)
            return Response({"error": "PDF generation failed."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        safe_title = re.sub(r"[^a-zA-Z0-9_-]", "_", title)[:40]
        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="novaa_{safe_title}.pdf"'
        return response


# ── NOVAA Send Email ──────────────────────────────────────────────────────────
class NovaaSendEmailView(APIView):
    """
    POST /api/ai/send-email/
    Body: { to_email, to_name, subject, body }
    Sends a single email drafted by NOVAA (student → teacher, etc.)
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        to_email = (request.data.get("to_email") or "").strip()
        to_name  = (request.data.get("to_name")  or "").strip()
        subject  = (request.data.get("subject")  or "Message from CampusEye").strip()
        body     = (request.data.get("body")      or "").strip()

        if not to_email or not body:
            return Response(
                {"error": "to_email and body are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        from attendance.services.novaa_action_executor import send_single_email
        result = send_single_email(
            sender_user_id=request.user.id,
            to_email=to_email,
            to_name=to_name,
            subject=subject,
            body=body,
        )

        if result["success"]:
            return Response(result, status=status.HTTP_200_OK)
        return Response(result, status=status.HTTP_502_BAD_GATEWAY)
